# -*- coding: utf-8 -*-
"""Exporta a Matriz de Criterios e Premissas (Etapa 4) para um arquivo Excel.

Gera tres abas:
  1. Dimensoes de Criterios -> Dimensao + Justificativa (com referencia)
  2. Criterios              -> # (continuo), Criterio, Dimensao
  3. Matriz de Criterios e Premissas -> matriz completa e preenchida,
     com a Fonte como link clicavel.

Fonte dos dados: templates/documento_apoio_construcao_modelo_conceitual.html (PREFILL.e4).
Texto revisado ortografica e gramaticalmente em portugues do Brasil.
"""

# flake8: noqa: E501
# Linhas longas sao intencionais: URLs de fontes e textos descritivos em
# portugues que nao devem ser quebrados para preservar a integridade dos dados.

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# --------------------------------------------------------------------------- #
# Dados                                                                        #
# --------------------------------------------------------------------------- #

# (dimensao, justificativa do agrupamento, com referencia bibliografica)
DIMENSOES: list[tuple[str, str]] = [
    (
        "Técnica",
        "Reúne os critérios de engenharia e operação que medem a necessidade "
        "física de intervenção: demanda de tráfego, saturação, congestionamento, "
        "estado de conservação, geometria e maturidade do projeto. O agrupamento "
        "justifica-se porque essas variáveis formam o eixo de desempenho "
        "operacional e de condição da infraestrutura, base de qualquer apreciação "
        "técnica de projetos rodoviários. Referências: Highway Capacity Manual "
        "(TRB, 2016); HDM-4 (Banco Mundial/PIARC); AASHTO — A Policy on Geometric "
        "Design of Highways and Streets (Green Book); UK DfT WebTAG.",
    ),
    (
        "Financeiro",
        "Agrupa os critérios de custo e de retorno monetário do investimento — "
        "Capex, Opex, relação benefício/custo, potencial de alavancagem privada e "
        "custo logístico. Separa-se da dimensão econômica por tratar do fluxo "
        "financeiro do próprio projeto e da eficiência do gasto público, conforme "
        "a análise custo-benefício consagrada na apreciação de projetos de "
        "transporte. Referências: US DOT Benefit-Cost Analysis Guidance; HM "
        "Treasury Green Book (2022); HEATCO (UE, 2006); World Bank PPP Reference "
        "Guide v3.",
    ),
    (
        "Econômica",
        "Concentra os impactos econômicos amplos (wider economic impacts): "
        "emprego, produção, competitividade, redução de tempos de viagem e "
        "desenvolvimento dos modos hidroviário e ferroviário. Distingue-se da "
        "dimensão financeira por medir os ganhos sistêmicos para a economia "
        "regional, e não o caixa do projeto. Referências: UK DfT WebTAG (wider "
        "economic impacts); HEATCO (UE, 2006); Macharis & Bernardini (2015, "
        "Transport Policy, 37:177-186).",
    ),
    (
        "Social",
        "Reúne os critérios de equidade e de acesso — redução de desigualdades "
        "regionais, população beneficiada, acessibilidade a serviços essenciais e "
        "atendimento a comunidades isoladas. O agrupamento justifica-se pela "
        "necessidade de avaliar os impactos distributivos e de acessibilidade de "
        "forma autônoma em relação à eficiência econômica. Referências: Geurs & "
        "van Wee (2004, Journal of Transport Geography, 12:127-140); UK DfT WebTAG "
        "(Social and Distributional Impacts).",
    ),
    (
        "Segurança",
        "Concentra os critérios de sinistralidade viária — acidentes com vítimas, "
        "acidentes com usuários vulneráveis, transporte de cargas perigosas e "
        "concentração de pontos críticos. Constitui dimensão própria por causa da "
        "natureza não compensatória do bem protegido (a vida humana) e do "
        "alinhamento ao Programa Vida no Trânsito e ao PNATRANS. Referências: iRAP "
        "(Star Rating); UK DfT WebTAG (accidents); PNATRANS (Brasil).",
    ),
    (
        "Ambiental",
        "Agrupa emissões de gases de efeito estufa, poluentes locais, eficiência "
        "energética, otimização da matriz modal, impacto sobre áreas protegidas e "
        "complexidade do licenciamento. O agrupamento atende à obrigação legal de "
        "considerar mudanças climáticas e à sustentabilidade como pilar do "
        "PLI-SP 2050. Referências: Lei Estadual 13.798/2009 (Política Estadual de "
        "Mudanças Climáticas); GHG Protocol; IPCC; Lei 6.938/1981; Decreto "
        "4.297/2002 (ZEE); Resolução CONAMA 237/1997.",
    ),
    (
        "Territorial",
        "Reúne a relação do projeto com o território e o uso do solo — conflitos "
        "com o tráfego urbano e portuário, integração intermodal, conexão "
        "inter-regional, aderência aos planos diretores e polos geradores de "
        "tráfego. Justifica-se pela necessidade de avaliar a inserção espacial do "
        "projeto e sua coerência com o ordenamento territorial. Referências: Lei "
        "10.257/2001 (Estatuto da Cidade); Highway Capacity Manual (TRB); ITE Trip "
        "Generation Manual; Macharis & Bernardini (2015).",
    ),
    (
        "Institucional",
        "Concentra a viabilidade institucional e jurídica — complexidade, prazo de "
        "implantação, pendências jurídicas, alinhamento aos planos, consenso entre "
        "atores e demanda social. O agrupamento decorre da abordagem multiator "
        "(MAMCA), que trata a legitimidade e a governança como dimensão própria de "
        "avaliação. Referências: Macharis & Bernardini (2015) — MAMCA; HM Treasury "
        "Green Book (modelo stage-gate); Lei 14.133/2021.",
    ),
    (
        "Risco",
        "Agrupa as incertezas que afetam a entrega do benefício — resiliência "
        "climática, risco de demanda, de execução, de desapropriação, de "
        "interdependência e socioambiental. Justifica-se pela evidência de viés de "
        "otimismo e de sobrecustos recorrentes em grandes projetos, o que exige o "
        "tratamento explícito do risco. Referências: Flyvbjerg (2009, Oxford "
        "Review of Economic Policy, 25(3):344-367); HM Treasury Green Book "
        "(optimism bias); World Bank ESF (2017); Convenção 169 da OIT; "
        "Decreto-Lei 3.365/1941.",
    ),
]

# Mapeia o simbolo da relacao para o rotulo exibido na ferramenta
REL = {
    "\u2191": "\u2191 Positiva",
    "\u2193": "\u2193 Negativa",
    "\u2195": "\u2195 Condicional",
}

# (dimensao, criterio, premissa, relacao, metricas, url_fonte, mandatorio)
CRITERIOS: list[tuple[str, str, str, str, str, str, str]] = [
    # ---- Técnica ----
    ("Técnica", "VDM — Volume Diário Médio",
     "O volume diário médio mede a demanda de tráfego já instalada no trecho. "
     "Quanto maior o VDM, maior a saturação e a pressão por intervenção, o que "
     "indica projetos que atendem a um número expressivo de usuários e geram "
     "benefício imediato.", "\u2191",
     "veículos/dia (VDM); modelo gravitacional",
     "https://www.gov.uk/guidance/transport-analysis-guidance-tag", "Não"),
    ("Técnica", "Nível de Serviço (NS) / Saturação",
     "O nível de serviço traduz a relação entre a demanda e a capacidade da via. "
     "Trechos que operam em níveis E ou F encontram-se congestionados ou "
     "sobre-saturados e exigem intervenção prioritária para restabelecer a "
     "fluidez e a segurança.", "\u2193",
     "NS de A a F; relação V/C no ano base e no futuro",
     "https://www.trb.org/Main/Blurbs/175169.aspx", "Não"),
    ("Técnica", "Congestionamento real (\"tempo lento\")",
     "A lentidão observada em plataformas de mapeamento em tempo real revela "
     "gargalos de capacidade que nem sempre aparecem nas contagens oficiais. "
     "Quanto maior o tempo adicional em relação ao fluxo livre, mais crítico é o "
     "trecho.", "\u2193",
     "minutos adicionais em relação ao fluxo livre (Google API)",
     "https://www.gov.uk/guidance/transport-analysis-guidance-tag", "Não"),
    ("Técnica", "Estado de conservação do pavimento",
     "A condição do pavimento afeta diretamente o custo operacional dos veículos, "
     "a segurança e a vida útil da via. Pavimentos degradados elevam os custos de "
     "manutenção e os acidentes, o que justifica priorizar a restauração.", "\u2193",
     "IRI/PCI; idade do pavimento",
     "https://documents.worldbank.org/en/publication/documents-reports/documentdetail/793271468171847482", "Não"),
    ("Técnica", "Deficiência geométrica (rampas, alças e raios)",
     "Rampas acentuadas, alças mal dimensionadas e raios de curva inadequados "
     "reduzem a capacidade e elevam o risco de acidentes. Corrigir essas "
     "deficiências geométricas melhora, ao mesmo tempo, a fluidez e a "
     "segurança.", "\u2193",
     "% de inclinação; raio de curva (m); velocidade (km/h)",
     "https://irap.org/rap-tools/infrastructure-ratings/star-ratings/", "Não"),
    ("Técnica", "Prontidão / maturidade do projeto",
     "Projetos com estudos e projeto executivo mais avançados podem ser licitados "
     "e entregues mais cedo, o que antecipa os benefícios à sociedade e reduz o "
     "risco de atraso na implantação.", "\u2191",
     "escala: ideia, estudo, projeto básico, projeto executivo",
     "https://www.gov.uk/government/publications/the-green-book-appraisal-and-evaluation-in-central-government", "Não"),
    ("Técnica", "Tráfego sazonal (fins de semana e feriados)",
     "Rodovias turísticas e de escoamento de safra sofrem picos sazonais que "
     "sobrecarregam a capacidade em fins de semana, feriados e épocas de "
     "colheita. A relação entre o VDM sazonal e o VDM médio dimensiona essa "
     "pressão adicional.", "\u2193",
     "razão entre VDM sazonal e VDM médio",
     "https://www.trb.org/Main/Blurbs/175169.aspx", "Não"),
    # ---- Financeiro ----
    ("Financeiro", "Capex (custo de investimento)",
     "O custo de investimento determina o volume de recursos públicos "
     "imobilizados. Projetos com menor Capex para um benefício equivalente "
     "liberam orçamento para mais intervenções e ampliam o alcance da "
     "carteira.", "\u2193",
     "R$ (custo total de investimento)",
     "https://www.transportation.gov/mission/office-secretary/office-policy/transportation-policy/benefit-cost-analysis-guidance", "Não"),
    ("Financeiro", "Opex (operação e manutenção)",
     "Os custos de operação e manutenção comprometem o orçamento ao longo de toda "
     "a vida útil do ativo. Projetos com menor Opex são financeiramente mais "
     "sustentáveis no longo prazo.", "\u2193",
     "R$/ano",
     "https://documents.worldbank.org/en/publication/documents-reports/documentdetail/793271468171847482", "Não"),
    ("Financeiro", "Relação benefício/custo",
     "A relação benefício/custo, o valor presente líquido (VPL) e a taxa interna "
     "de retorno (TIR) sintetizam o retorno por real investido. Quanto maior o "
     "retorno, maior a eficiência da aplicação dos recursos públicos.", "\u2191",
     "B/C; VPL; TIR",
     "https://www.gov.uk/government/publications/the-green-book-appraisal-and-evaluation-in-central-government", "Não"),
    ("Financeiro", "Potencial de financiamento privado / concessão",
     "Projetos com potencial de concessão ou de parceria público-privada permitem "
     "alavancar capital privado, o que preserva o caixa do Estado e amplia a "
     "capacidade de investimento.", "\u2191",
     "% de alavancagem por PPP/concessão",
     "https://ppp.worldbank.org/sites/default/files/2024-08/PPP%20Reference%20Guide%20Version%203.pdf", "Não"),
    ("Financeiro", "Custos logísticos diferenciados (porto/corredor)",
     "Diferenciais de custo logístico entre corredores e portos — por exemplo, o "
     "sobrecusto por tonelada de soja escoada por Santos — afetam a "
     "competitividade. Reduzir esses custos é critério relevante de "
     "priorização.", "\u2193",
     "US$/tonelada (ex.: +US$ 5/t de soja por Santos)",
     "https://trimis.ec.europa.eu/project/developing-harmonised-european-approaches-transport-costing-and-project-assessment", "Não"),
    ("Financeiro", "Benefício social na priorização",
     "Além do retorno financeiro privado, deve-se considerar o valor presente "
     "líquido social, que incorpora externalidades positivas e negativas e alinha "
     "a seleção ao maior benefício agregado para a sociedade.", "\u2191",
     "VPL social; B/C social",
     "https://www.gov.uk/government/publications/the-green-book-appraisal-and-evaluation-in-central-government", "Não"),
    # ---- Econômica ----
    ("Econômica", "Empregos e produção envolvidos",
     "Projetos que servem a uma base produtiva maior, medida em empregos e em "
     "valor de produção, geram impacto econômico regional mais expressivo e devem "
     "ter prioridade na carteira.", "\u2191",
     "número de empregos; R$/ano de produção",
     "https://www.gov.uk/guidance/transport-analysis-guidance-tag", "Não"),
    ("Econômica", "Competitividade da produção paulista",
     "A redução do custo logístico aumenta a competitividade da produção paulista "
     "nos mercados interno e externo, o que ajuda a manter e a ampliar a "
     "participação de mercado do Estado.", "\u2191",
     "participação de mercado; custo logístico total",
     "https://www.gov.uk/guidance/transport-analysis-guidance-tag", "Não"),
    ("Econômica", "Redução dos tempos de viagem",
     "A economia de tempo de viagem de pessoas e cargas converte-se diretamente "
     "em ganho de produtividade e em redução de custos, sendo um dos principais "
     "benefícios mensuráveis dos projetos de transporte.", "\u2193",
     "min/km; tempo total origem-destino",
     "https://trimis.ec.europa.eu/project/developing-harmonised-european-approaches-transport-costing-and-project-assessment", "Não"),
    ("Econômica", "Indução de produção e emprego regional",
     "Infraestrutura nova pode induzir o surgimento de produção e de emprego em "
     "regiões-alvo, atuando como vetor de desenvolvimento e de desconcentração "
     "econômica do Estado.", "\u2191",
     "empregos e PIB induzidos por região",
     "https://www.gov.uk/guidance/transport-analysis-guidance-tag", "Não"),
    ("Econômica", "Atendimento a cargas sem alternativa eficiente",
     "Há cargas hoje sem alternativa de transporte eficiente, presas a modais "
     "inadequados. Destravar essa demanda reprimida gera ganho econômico "
     "expressivo e amplia a eficiência da rede.", "\u2191",
     "toneladas/ano de carga atendida",
     "https://trid.trb.org/view/1334470", "Não"),
    ("Econômica", "Suporte a cadeias estratégicas (Pré-Sal, agro, sucroenergético)",
     "Cadeias-chave do Estado — Pré-Sal, agronegócio e sucroenergético — dependem "
     "de logística adequada. Viabilizar o transporte dessas cadeias fortalece "
     "setores estratégicos da economia paulista.", "\u2191",
     "toneladas/ano movimentadas das cadeias-alvo",
     "https://trid.trb.org/view/1334470", "Não"),
    ("Econômica", "Participação da hidrovia na matriz de transporte",
     "Ampliar a participação da hidrovia Tietê-Paraná na matriz de transporte "
     "reduz custos e emissões, ao aproveitar um modal de alta eficiência ainda "
     "subutilizado no Estado.", "\u2191",
     "% de TKM hidroviário",
     "https://trid.trb.org/view/1334470", "Não"),
    ("Econômica", "Desenvolvimento ferroviário estadual",
     "Viabilizar ferrovias e short lines reduz a ociosidade da malha existente e "
     "desloca carga do modal rodoviário para um modal mais eficiente em longas "
     "distâncias.", "\u2191",
     "km de malha ativa; TKM ferroviário",
     "https://trid.trb.org/view/1334470", "Não"),
    # ---- Social ----
    ("Social", "Redução de desigualdades regionais",
     "Direcionar o investimento preferencialmente a regiões menos desenvolvidas "
     "reduz as desigualdades territoriais. É critério obrigatório por orientação "
     "do PPT (slide 3) e da política de desenvolvimento regional.", "\u2191",
     "IDH regional; PIB per capita regional",
     "https://www.gov.uk/guidance/transport-analysis-guidance-tag", "Sim"),
    ("Social", "População beneficiada",
     "Quanto maior a população na área de influência do projeto, maior o retorno "
     "social do investimento, pois um número maior de pessoas usufrui da "
     "melhoria.", "\u2191",
     "habitantes na área de influência",
     "https://www.gov.uk/guidance/transport-analysis-guidance-tag", "Não"),
    ("Social", "Equidade no acesso ao transporte",
     "Projetos que ampliam o acesso ao transporte para populações hoje "
     "subatendidas promovem equidade e inclusão, ao reduzir barreiras de "
     "mobilidade.", "\u2191",
     "% da população atendida por serviço público",
     "https://research.utwente.nl/en/publications/accessibility-evaluation-of-land-use-and-transport-strategies-rev", "Não"),
    ("Social", "Acesso a serviços essenciais (saúde e educação)",
     "Conectar a população a unidades básicas de saúde, hospitais e escolas é "
     "função social essencial do transporte. Projetos que melhoram esse acesso "
     "têm alto valor social.", "\u2191",
     "população com acesso a serviço essencial via projeto",
     "https://research.utwente.nl/en/publications/accessibility-evaluation-of-land-use-and-transport-strategies-rev", "Não"),
    ("Social", "Acessibilidade a polos (portos e aeroportos)",
     "Melhorar o acesso terrestre a polos logísticos — portos de Santos e de São "
     "Sebastião e Aeroporto de Viracopos — amplia a eficiência da cadeia e o "
     "retorno dos investimentos já realizados nesses polos.", "\u2191",
     "tempo médio de acesso; NS do acesso",
     "https://research.utwente.nl/en/publications/accessibility-evaluation-of-land-use-and-transport-strategies-rev", "Não"),
    ("Social", "Atendimento a comunidades isoladas",
     "Conectar comunidades hoje isoladas reduz o isolamento territorial e amplia "
     "o acesso dessas populações a oportunidades, mercados e serviços "
     "públicos.", "\u2191",
     "número de comunidades conectadas",
     "https://research.utwente.nl/en/publications/accessibility-evaluation-of-land-use-and-transport-strategies-rev", "Não"),
    # ---- Segurança ----
    ("Segurança", "Acidentes com vítimas (gravidade)",
     "A redução de mortes e de feridos graves é prioridade de saúde pública. "
     "Trechos com maior número de óbitos e de feridos graves por quilômetro devem "
     "ser tratados prioritariamente.", "\u2193",
     "óbitos e feridos graves por km (InfoSiga)",
     "https://irap.org/rap-tools/infrastructure-ratings/star-ratings/", "Não"),
    ("Segurança", "Acidentes com usuários vulneráveis",
     "Pedestres, ciclistas e motociclistas são os usuários mais expostos da via. "
     "Projetos que reduzem os acidentes com esses grupos têm alto impacto na "
     "segurança viária.", "\u2193",
     "ocorrências/ano com usuários vulneráveis (InfoSiga)",
     "https://irap.org/rap-tools/infrastructure-ratings/star-ratings/", "Não"),
    ("Segurança", "Transporte de cargas perigosas",
     "O transporte de produtos perigosos exige mitigação de risco antes, durante "
     "e após eventuais acidentes, dada a gravidade potencial de vazamentos, "
     "incêndios e explosões.", "\u2193",
     "ocorrências/ano; severidade",
     "https://irap.org/rap-tools/infrastructure-ratings/star-ratings/", "Não"),
    ("Segurança", "Concentração de pontos críticos (\"black spots\")",
     "Trechos com concentração de acidentes (pontos críticos) apresentam alto "
     "retorno em vidas salvas por real investido e justificam intervenção "
     "prioritária.", "\u2193",
     "densidade de acidentes por km",
     "https://irap.org/rap-tools/infrastructure-ratings/star-ratings/", "Não"),
    # ---- Ambiental ----
    ("Ambiental", "Redução de emissões de gases de efeito estufa (GEE)",
     "A redução das emissões de gases de efeito estufa é diretriz da Política "
     "Estadual de Mudanças Climáticas (Lei 13.798/2009). É critério obrigatório e "
     "mensurável em toneladas de CO\u2082 equivalente evitadas.", "\u2193",
     "tCO\u2082eq/ano evitadas",
     "https://www.al.sp.gov.br/repositorio/legislacao/lei/2009/lei-13798-09.11.2009.html", "Sim"),
    ("Ambiental", "Redução de poluentes locais",
     "A redução de poluentes locais, como óxidos de nitrogênio e material "
     "particulado, melhora a qualidade do ar e a saúde pública nas áreas "
     "atravessadas pelo projeto.", "\u2193",
     "g de NOx/km; material particulado",
     "https://trimis.ec.europa.eu/project/developing-harmonised-european-approaches-transport-costing-and-project-assessment", "Não"),
    ("Ambiental", "Eficiência energética",
     "Projetos que reduzem o consumo de energia por tonelada-quilômetro "
     "transportada diminuem custos e emissões, o que contribui para a "
     "sustentabilidade do sistema de transporte.", "\u2191",
     "MJ/(t.km) evitados",
     "https://www.ipcc.ch/", "Não"),
    ("Ambiental", "Otimização da matriz modal",
     "Estimular a migração de carga do modal rodoviário para os modais "
     "ferroviário e hidroviário, mais sustentáveis, é objetivo ambiental central "
     "do PLI-SP 2050.", "\u2191",
     "% de migração rodoviário para ferroviário/hidroviário",
     "https://trid.trb.org/view/1334470", "Não"),
    ("Ambiental", "Impacto sobre áreas sensíveis ou protegidas",
     "Projetos que afetam áreas de preservação permanente, unidades de "
     "conservação ou a zona costeira geram passivo ambiental. Evitar ou minimizar "
     "esse impacto é critério obrigatório, inclusive quanto ao conflito com o "
     "Zoneamento Ecológico-Econômico (ZEE).", "\u2193",
     "área protegida afetada; conflito com o ZEE",
     "https://www.planalto.gov.br/ccivil_03/decreto/2002/d4297.htm", "Sim"),
    ("Ambiental", "Complexidade do licenciamento ambiental",
     "A complexidade e o prazo do licenciamento ambiental podem atrasar ou "
     "inviabilizar a obra. Projetos com licenciamento mais simples apresentam "
     "menor risco de execução.", "\u2193",
     "grau e prazo do licenciamento",
     "https://conama.mma.gov.br/?option=com_sisconama&task=arquivo.download&id=237", "Não"),
    # ---- Territorial ----
    ("Territorial", "Conflito com o tráfego urbano e conurbações",
     "Em conurbações, o tráfego de passagem disputa espaço com o tráfego local e "
     "gera conflito urbano-regional. Projetos que reduzem essa interferência "
     "melhoram a fluidez e a qualidade de vida urbana.", "\u2193",
     "horas de pico; índice de conflito",
     "https://www.trb.org/Main/Blurbs/175169.aspx", "Não"),
    ("Territorial", "Conflito com o tráfego urbano portuário",
     "Nas cidades portuárias, o tráfego de cargas disputa o viário urbano com a "
     "população. Reduzir o conflito urbano-portuário aumenta a eficiência "
     "logística e reduz as externalidades negativas.", "\u2193",
     "índice de interferência urbana",
     "https://trid.trb.org/view/1334470", "Não"),
    ("Territorial", "Integração intermodal",
     "Projetos que conectam diferentes modais — portos, ferrovias e aeroportos — "
     "ampliam o benefício logístico do conjunto da rede, ao viabilizar cadeias de "
     "transporte integradas.", "\u2191",
     "número de conexões modais",
     "https://trid.trb.org/view/1334470", "Não"),
    ("Territorial", "Conexão inter-regional e vazios logísticos",
     "Ligações que integram regiões hoje mal conectadas reduzem vazios logísticos "
     "e ampliam a coesão territorial do Estado.", "\u2191",
     "novas ligações inter-regionais; frequência",
     "https://research.utwente.nl/en/publications/accessibility-evaluation-of-land-use-and-transport-strategies-rev", "Não"),
    ("Territorial", "Aderência aos planos diretores municipais",
     "A compatibilidade do projeto com os planos diretores municipais e com o uso "
     "do solo planejado evita conflitos futuros e confere segurança jurídica à "
     "implantação.", "\u2191",
     "grau de compatibilidade",
     "https://www.planalto.gov.br/ccivil_03/leis/leis_2001/l10257.htm", "Não"),
    ("Territorial", "Polos atratores e geradores de tráfego",
     "Polos atratores e geradores de tráfego — terminais, centros logísticos e "
     "grandes equipamentos urbanos — ampliam a demanda sobre a via. "
     "Considerá-los dimensiona corretamente a necessidade de capacidade.", "\u2191",
     "entradas e saídas de terminais por dia",
     "https://www.trb.org/Main/Blurbs/175169.aspx", "Não"),
    # ---- Institucional ----
    ("Institucional", "Nível de complexidade",
     "Projetos menos complexos do ponto de vista técnico e institucional tendem a "
     "ser executados com mais rapidez e menor risco, entregando resultado em "
     "prazo mais curto. É critério obrigatório (PPT, slide 3).", "\u2193",
     "escala qualitativa de 1 a 5",
     "https://www.gov.uk/government/publications/the-green-book-appraisal-and-evaluation-in-central-government", "Sim"),
    ("Institucional", "Prazo para implantação",
     "Quanto menor o prazo até a entrada em operação, mais cedo os benefícios são "
     "gerados. O prazo de implantação é critério obrigatório de priorização "
     "(PPT, slide 3).", "\u2193",
     "meses até a operação",
     "https://www.gov.uk/government/publications/the-green-book-appraisal-and-evaluation-in-central-government", "Sim"),
    ("Institucional", "Pendências jurídicas e jurisdicionais",
     "A ausência de impedimentos jurídicos e jurisdicionais é condição para a "
     "execução. Pendências elevam o risco de paralisação e constituem critério "
     "obrigatório (PPT, slide 3).", "\u2193",
     "número de pendências; grau de risco jurídico",
     "https://www.planalto.gov.br/ccivil_03/_ato2019-2022/2021/lei/l14133.htm", "Sim"),
    ("Institucional", "Alinhamento aos planos (PPA, PEF, PAN, PNLT)",
     "A coerência do projeto com os planos vigentes — PPA, PEF, PAN e PNLT — "
     "garante legitimidade orçamentária e institucional à priorização.", "\u2191",
     "grau de aderência aos planos",
     "https://trid.trb.org/view/1334470", "Não"),
    ("Institucional", "Consenso e apoio institucional dos atores",
     "A convergência entre os atores institucionais envolvidos facilita a "
     "execução e reduz o risco de bloqueios políticos ou administrativos.", "\u2191",
     "grau de consenso (atores do slide 58)",
     "https://trid.trb.org/view/1334470", "Não"),
    ("Institucional", "Demanda social e contribuições recebidas",
     "O volume de contribuições e de manifestações sociais recebidas confere "
     "legitimidade participativa à priorização e indica o apoio da sociedade ao "
     "projeto.", "\u2191",
     "número de menções e contribuições",
     "https://trid.trb.org/view/1334470", "Não"),
    # ---- Risco ----
    ("Risco", "Resiliência climática (Blue Spot)",
     "Projetos que aumentam a resiliência da rede a eventos climáticos extremos "
     "(pontos de alagamento, ou Blue Spots) reduzem o risco de interrupção do "
     "tráfego e os custos de recuperação.", "\u2191",
     "% da rede resiliente; nº de Blue Spots mitigados",
     "https://www.ipcc.ch/", "Não"),
    ("Risco", "Risco de demanda (incerteza de projeção)",
     "Projeções de demanda incertas elevam o risco do investimento. Quanto maior "
     "a variância entre os cenários, maior a chance de o benefício previsto não "
     "se concretizar.", "\u2193",
     "variância entre cenários de demanda",
     "https://academic.oup.com/oxrep/article-abstract/25/3/344/424009", "Não"),
    ("Risco", "Risco de execução (atrasos e sobrecustos)",
     "O histórico de desvios de prazo e de custo — viés de otimismo — sinaliza "
     "maior risco de execução. Projetos com esse histórico devem ser penalizados "
     "na priorização.", "\u2193",
     "desvio histórico de prazo e de custo",
     "https://academic.oup.com/oxrep/article-abstract/25/3/344/424009", "Não"),
    ("Risco", "Risco de desapropriação e interferências",
     "Desapropriações e interferências com redes de infraestrutura são causas "
     "frequentes de atraso e de sobrecusto. Quanto maior o número de "
     "desapropriações e de interferências, maior o risco.", "\u2193",
     "número de desapropriações; interferências",
     "https://www.planalto.gov.br/ccivil_03/decreto-lei/del3365.htm", "Não"),
    ("Risco", "Dependência de pré-requisitos (interdependência)",
     "Projetos cujo benefício depende da execução prévia de outros (predecessores) "
     "carregam risco maior, pois ficam reféns do cronograma dos projetos dos "
     "quais dependem.", "\u2193",
     "número de projetos predecessores",
     "https://www.gov.uk/government/publications/the-green-book-appraisal-and-evaluation-in-central-government", "Não"),
    ("Risco", "Risco socioambiental (comunidades tradicionais)",
     "A presença de comunidades tradicionais, como quilombolas e indígenas, exige "
     "consulta e cuidado redobrados. O risco socioambiental, se não for tratado, "
     "pode paralisar o projeto.", "\u2193",
     "presença e grau de conflito socioambiental",
     "https://www.ilo.org/pt-pt/resource/other/convencao-no-169-da-oit-sobre-povos-indigenas-e-tribais-traduzida-para", "Não"),
]

# --------------------------------------------------------------------------- #
# Estilos                                                                      #
# --------------------------------------------------------------------------- #

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
MAND_FILL = PatternFill("solid", fgColor="FCE4D6")
MAND_FONT = Font(bold=True, color="C00000")
DIM_FONT = Font(bold=True, color="1F4E78")
LINK_FONT = Font(color="0563C1", underline="single")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_header(ws, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"


def apply_grid(ws, nrows: int, ncols: int) -> None:
    for r in range(2, nrows + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            if cell.alignment is None or cell.alignment.wrap_text is None:
                cell.alignment = WRAP_TOP


# --------------------------------------------------------------------------- #
# Construção das abas                                                          #
# --------------------------------------------------------------------------- #

def _require_active_sheet(wb: Workbook) -> Worksheet:
    ws = wb.active
    if ws is None:
        raise RuntimeError("Workbook sem aba ativa.")
    return ws


def aba_dimensoes(wb: Workbook) -> None:
    ws = _require_active_sheet(wb)
    ws.title = "Dimensões de Critérios"
    ws.append(["#", "Dimensão", "Justificativa (com referência)"])
    for i, (nome, just) in enumerate(DIMENSOES, start=1):
        ws.append([i, nome, just])
    style_header(ws, 3)
    apply_grid(ws, len(DIMENSOES) + 1, 3)
    for idx, w in enumerate([5, 16, 118], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    for r in range(2, len(DIMENSOES) + 2):
        ws.cell(row=r, column=1).alignment = CENTER
        ws.cell(row=r, column=2).font = DIM_FONT
        ws.cell(row=r, column=2).alignment = WRAP_TOP
        ws.cell(row=r, column=3).alignment = WRAP_TOP
        ws.row_dimensions[r].height = 104


def aba_criterios(wb: Workbook) -> None:
    ws = wb.create_sheet("Critérios")
    ws.append(["#", "Critério", "Dimensão"])
    for i, c in enumerate(CRITERIOS, start=1):
        ws.append([i, c[1], c[0]])
    style_header(ws, 3)
    apply_grid(ws, len(CRITERIOS) + 1, 3)
    for idx, w in enumerate([6, 56, 18], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    for r in range(2, len(CRITERIOS) + 2):
        ws.cell(row=r, column=1).alignment = CENTER
        ws.cell(row=r, column=2).alignment = WRAP_TOP
        ws.cell(row=r, column=3).font = DIM_FONT
        ws.cell(row=r, column=3).alignment = WRAP_TOP


def aba_matriz(wb: Workbook) -> None:
    ws = wb.create_sheet("Matriz de Critérios e Premissas")
    headers = ["Dimensão", "Critério", "Premissa", "Relação",
               "Métricas", "Fonte", "Mandatório"]
    ws.append(headers)
    for c in CRITERIOS:
        ws.append([c[0], c[1], c[2], REL.get(c[3], c[3]), c[4], c[5], c[6]])
    style_header(ws, len(headers))
    apply_grid(ws, len(CRITERIOS) + 1, len(headers))
    for idx, w in enumerate([13, 32, 56, 13, 26, 42, 12], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    for r in range(2, len(CRITERIOS) + 2):
        ws.cell(row=r, column=1).font = DIM_FONT
        ws.cell(row=r, column=1).alignment = WRAP_TOP
        ws.cell(row=r, column=2).alignment = WRAP_TOP
        ws.cell(row=r, column=3).alignment = WRAP_TOP
        ws.cell(row=r, column=4).alignment = CENTER
        ws.cell(row=r, column=5).alignment = WRAP_TOP
        link = ws.cell(row=r, column=6)
        link.hyperlink = link.value
        link.font = LINK_FONT
        link.alignment = Alignment(wrap_text=True, vertical="top")
        mand = ws.cell(row=r, column=7)
        mand.alignment = CENTER
        if mand.value == "Sim":
            mand.fill = MAND_FILL
            mand.font = MAND_FONT
        ws.row_dimensions[r].height = 66


def main() -> None:
    wb = Workbook()
    aba_dimensoes(wb)
    aba_criterios(wb)
    aba_matriz(wb)
    base = Path(__file__).resolve().parents[1] / "data"
    out = base / "Matriz_Criterios_Premissas_PLI-SP.xlsx"
    try:
        wb.save(out)
    except PermissionError:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        out = base / f"Matriz_Criterios_Premissas_PLI-SP_{stamp}.xlsx"
        wb.save(out)
        print(
            "Aviso: o arquivo padrao estava aberto/bloqueado; "
            "salvo com nome alternativo."
        )
    mand = sum(1 for c in CRITERIOS if c[6] == "Sim")
    print(f"Arquivo gerado: {out}")
    print(f"  - Dimensões: {len(DIMENSOES)}")
    print(f"  - Critérios: {len(CRITERIOS)}")
    print(f"  - Mandatórios: {mand} | Não mandatórios: {len(CRITERIOS) - mand}")


if __name__ == "__main__":
    main()
