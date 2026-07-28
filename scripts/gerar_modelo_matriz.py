# Gera o arquivo-modelo da Matriz de Premissas e Critérios (formato congelado)
# a partir da estrutura de data/Matriz_Criterios_Premissas_PLI-SP_FASES_V2_20260715_ATUALIZADA.xlsx
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "data" / "modelos" / "Modelo_Matriz_Criterios_Premissas_SLT.xlsx"
OUT.parent.mkdir(parents=True, exist_ok=True)

HEAD_FILL = PatternFill("solid", fgColor="1F4E5F")
HEAD_FONT = Font(bold=True, color="FFFFFF", size=11)
EX_FILL = PatternFill("solid", fgColor="FFF3CD")
EX_FONT = Font(italic=True, color="7A6000")
THIN = Border(*[Side(style="thin", color="B0B0B0")] * 4)
WRAP = Alignment(wrap_text=True, vertical="top")

wb = Workbook()


def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.border = THIN
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")


def add_sheet(title, headers, widths, examples=None, presets=None, freeze="A2", n_blank=200):
    ws = wb.create_sheet(title)
    ws.append(headers)
    style_header(ws, len(headers))
    # Linhas pré-populadas (dados canônicos do SLT/PLI-SP). Não são exemplos e não devem ser apagadas.
    for row in presets or []:
        ws.append(row)
        r = ws.max_row
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN
            cell.alignment = WRAP
    # Linhas de exemplo (destacadas em amarelo) que o usuário deve remover antes de enviar.
    for ex in examples or []:
        ws.append(ex)
        r = ws.max_row
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = EX_FILL
            cell.font = EX_FONT
            cell.border = THIN
            cell.alignment = WRAP
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    for r in range(ws.max_row + 1, ws.max_row + 1 + n_blank):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN
            cell.alignment = WRAP
    ws.freeze_panes = freeze
    ws.sheet_properties.tabColor = "1F4E5F"
    return ws


# ---------------- Aba de instruções ----------------
inst = wb.active
inst.title = "Instruções"
inst.sheet_properties.tabColor = "C0392B"
linhas = [
    ("MODELO — MATRIZ DE PREMISSAS E CRITÉRIOS (SLT / PLI-SP)", ""),
    ("", ""),
    ("Como preencher", ""),
    ("1.", "Não altere os nomes das abas nem os cabeçalhos das colunas: o sistema depende deles para interpretar o arquivo."),
    ("2.", "As linhas destacadas em amarelo são EXEMPLOS de preenchimento. Apague-as antes de enviar o arquivo."),
    ("3.", "A aba 'Dimensões de Critérios' já vem com as 9 dimensões canônicas do SLT/PLI-SP pré-preenchidas. Ajuste apenas se a rodada exigir dimensões distintas; caso contrário, mantenha-as."),
    ("4.", "Preencha a aba 'Critérios' com todos os critérios utilizados e a aba 'Matriz Crit Premissas v2' com a matriz completa (uma linha por premissa)."),
    ("5.", "Coluna 'Etapa' (Matriz Crit Premissas v2): use apenas 'Elegibilidade', 'Favorabilidade' ou 'Ajuste de Prioridade'."),
    ("6.", "Coluna 'Relação': use '↓ Negativa' quando valores maiores pioram a avaliação e '↑ Positiva' quando valores maiores melhoram."),
    ("7.", "Coluna 'Mandatório': use apenas 'Sim' ou 'Não'."),
    ("8.", "Coluna 'Classificação': ex.: 'restrição legal', 'risco inicial / restrição condicionada', 'risco espacial', 'favorabilidade'."),
    ("9.", "Não mescle células, não insira colunas novas e não use fórmulas nas células de dados."),
    ("10.", "Salve o arquivo em formato .xlsx e anexe-o na seção 'Matriz de premissas e critérios' do cadastro da hierarquização."),
    ("", ""),
    ("Abas obrigatórias", "Dimensões de Critérios · Critérios · Matriz Crit Premissas v2"),
]
for a, b in linhas:
    inst.append([a, b])
inst.column_dimensions["A"].width = 8
inst.column_dimensions["B"].width = 130
inst["A1"].font = Font(bold=True, size=14, color="1F4E5F")
inst["A3"].font = Font(bold=True, size=12)
for r in range(1, inst.max_row + 1):
    inst.cell(row=r, column=2).alignment = WRAP

# ---------------- Dimensões ----------------
# 9 dimensões canônicas do SLT/PLI-SP (data/matriz-criterios-premissas.json).
DIMENSOES_CANONICAS = [
    (1, "Técnica",
     "Agrupa critérios de engenharia e operação que medem a necessidade física de intervenção (demanda, saturação, congestionamento, conservação, geometria e maturidade). Ref.: Highway Capacity Manual (TRB, 2016); HDM-4 (Banco Mundial/PIARC); AASHTO A Policy on Geometric Design (Green Book); UK DfT WebTAG."),
    (2, "Financeiro",
     "Reúne os critérios de custo e retorno monetário do investimento (Capex, Opex, B/C, alavancagem privada e custo logístico). Separa a dimensão estritamente financeira da econômica, conforme a análise custo-benefício padrão em apreciação de transportes. Ref.: US DOT Benefit-Cost Analysis Guidance; HM Treasury Green Book (2022); HEATCO (UE, 2006); World Bank PPP Reference Guide v3."),
    (3, "Econômica",
     "Captura os impactos econômicos amplos (wider economic impacts): emprego, produção, competitividade, tempos de viagem e desenvolvimento modal. Distingue-se do Financeiro por medir externalidades e ganhos sistêmicos para a economia regional. Ref.: UK DfT WebTAG (wider economic impacts); HEATCO (UE, 2006); Macharis & Bernardini (2015, Transport Policy 37:177-186)."),
    (4, "Social",
     "Agrupa critérios de equidade e acesso (desigualdades regionais, população beneficiada, acessibilidade a serviços essenciais e comunidades isoladas). Ref.: Geurs & van Wee (2004, J. Transport Geography 12:127-140); UK DfT WebTAG (Social & Distributional Impacts)."),
    (5, "Segurança",
     "Concentra a sinistralidade viária (acidentes com vítimas, usuários vulneráveis, cargas perigosas e pontos críticos). O agrupamento próprio decorre da natureza não-compensatória do critério (vidas humanas). Ref.: iRAP Star Ratings; UK DfT WebTAG (accidents); PNATRANS (Brasil)."),
    (6, "Ambiental",
     "Reúne emissões de GEE, poluentes locais, eficiência energética, matriz modal, impacto sobre áreas protegidas e licenciamento. Ref.: Lei Estadual 13.798/2009 (PEMC); GHG Protocol; IPCC; Lei 6.938/1981; Decreto 4.297/2002 (ZEE); Resolução CONAMA 237/1997."),
    (7, "Territorial",
     "Agrupa a relação do projeto com o território e o uso do solo (conflitos urbanos, integração intermodal, conexão inter-regional, planos diretores e polos geradores). Ref.: Lei 10.257/2001 (Estatuto da Cidade); Highway Capacity Manual (TRB); ITE Trip Generation; Macharis & Bernardini (2015)."),
    (8, "Institucional",
     "Reúne a viabilidade institucional e jurídica (complexidade, prazo, pendências jurídicas, alinhamento a planos, consenso entre atores e demanda social). Ref.: Macharis & Bernardini (2015) - MAMCA; HM Treasury Green Book (stage-gate); Lei 14.133/2021."),
    (9, "Risco",
     "Concentra as incertezas que afetam a entrega do benefício (resiliência climática, risco de demanda, de execução, de desapropriação, de interdependência e socioambiental). Ref.: Flyvbjerg (2009, Oxford Review of Economic Policy 25(3):344-367); HM Treasury Green Book (optimism bias); World Bank ESF (2017); Convenção 169 OIT; Decreto-Lei 3.365/1941."),
]

add_sheet(
    "Dimensões de Critérios",
    ["#", "Dimensão", "Justificativa (com referência)"],
    [6, 25, 100],
    presets=[list(d) for d in DIMENSOES_CANONICAS],
)

# ---------------- Critérios ----------------
add_sheet(
    "Critérios",
    ["#", "Critério", "Dimensão"],
    [6, 70, 25],
    examples=[[1, "EXEMPLO — Sobreposição com Unidade de Conservação de Proteção Integral", "Ambiental"]],
)

# ---------------- Matriz ----------------
# Ordem das colunas: Etapa fica imediatamente antes de Classificação.
mat_headers = [
    "Dimensão", "Critério", "Premissa", "Etapa", "Classificação", "Dado",
    "Variável", "Unidade de medida", "Operador", "Relação", "Métricas",
    "Fonte", "Mandatório",
]
ws_mat = add_sheet(
    "Matriz Crit Premissas v2",
    mat_headers,
    [16, 45, 60, 22, 30, 45, 45, 26, 45, 14, 45, 45, 12],
    examples=[[
        "Ambiental",
        "EXEMPLO — Sobreposição com Unidade de Conservação de Proteção Integral (estadual)",
        "As unidades estaduais do grupo de Proteção Integral restringem a implantação de empreendimentos.",
        "Elegibilidade",
        "risco inicial / restrição condicionada",
        "Polígonos oficiais das Unidades de Conservação estaduais",
        "Interseção da demanda com cada feição de UC de Proteção Integral",
        "presença/ausência (0/1)",
        "União temática por Identity, preservando os atributos da UC",
        "↓ Negativa",
        "0 sem interseção; 1 com interseção",
        "DataGeo/IDEA-SP e Fundação Florestal; Lei Federal nº 9.985/2000",
        "Sim",
    ]],
)
# Colunas na nova ordem: A=Dimensão, B=Critério, C=Premissa, D=Etapa, E=Classificação,
# F=Dado, G=Variável, H=Unidade de medida, I=Operador, J=Relação, K=Métricas, L=Fonte, M=Mandatório.
dv_etapa = DataValidation(
    type="list",
    formula1='"Elegibilidade,Favorabilidade,Ajuste de Prioridade"',
    allow_blank=True,
)
dv_rel = DataValidation(type="list", formula1='"↓ Negativa,↑ Positiva"', allow_blank=True)
dv_mand = DataValidation(type="list", formula1='"Sim,Não"', allow_blank=True)
for dv, col in ((dv_etapa, "D"), (dv_rel, "J"), (dv_mand, "M")):
    ws_mat.add_data_validation(dv)
    dv.add(f"{col}2:{col}500")

wb.save(OUT)
print("OK:", OUT)
