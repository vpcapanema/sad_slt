# -*- coding: utf-8 -*-
"""Gera planilha-modelo para etapa conceitual da análise AHP (upload no SLT).

Aba principal: «Tabela de Premissas e Critérios» (colunas A–G)
  Dimensao | Criterio | Premissa | Relacao | Metricas | Fonte | Mandatorio

- Listas suspensas: Dimensao, Relacao, Mandatorio
- Colunas H em diante ocultas
- Aba «Exemplo de preenchimento» com instrucoes e linhas reais (PLI-SP)
"""

from __future__ import annotations

import importlib.util
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "data" / "Modelo_Matriz_Criterios_Premissas_Analise.xlsx"
OUT_CSV = ROOT / "data" / "Modelo_Matriz_Criterios_Premissas_Analise.csv"

# Reutiliza dimensoes, criterios e rotulos de relacao do export oficial
_spec = importlib.util.spec_from_file_location(
    "export_matriz_excel",
    Path(__file__).parent / "export_matriz_excel.py",
)
_mod = importlib.util.module_from_spec(_spec)
assert _spec.loader is not None
_spec.loader.exec_module(_mod)

DIMENSOES = _mod.DIMENSOES
CRITERIOS = _mod.CRITERIOS
REL = _mod.REL

HEADERS = [
    "Dimensao",
    "Criterio",
    "Premissa",
    "Relacao",
    "Metricas",
    "Fonte",
    "Mandatorio",
]

RELACOES = list(REL.values())
MANDATORIO = ["Sim", "Não"]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
INSTR_FILL = PatternFill("solid", fgColor="E8F4FC")
INSTR_FONT = Font(size=10, color="1F4E78")
MAND_FILL = PatternFill("solid", fgColor="FCE4D6")
MAND_FONT = Font(bold=True, color="C00000")
DIM_FONT = Font(bold=True, color="1F4E78")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

MAX_DATA_ROW = 200
HIDDEN_FROM_COL = 8  # H


def _style_header(ws, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"


def _hide_cols_from_h(ws) -> None:
    for col in range(HIDDEN_FROM_COL, 40):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].hidden = True


def _col_widths(ws) -> None:
    for idx, w in enumerate([16, 34, 52, 14, 28, 38, 12], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w


def _apply_grid(ws, start_row: int, end_row: int, ncols: int = 7) -> None:
    for r in range(start_row, end_row + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            if cell.alignment is None:
                cell.alignment = WRAP_TOP
        ws.row_dimensions[r].height = 48


def _aba_listas(wb: Workbook):
    ws = wb.create_sheet("_Listas")
    ws.append(["Dimensao", "Relacao", "Mandatorio"])
    n = max(len(DIMENSOES), len(RELACOES), len(MANDATORIO))
    for i in range(n):
        dim = DIMENSOES[i][0] if i < len(DIMENSOES) else None
        rel = RELACOES[i] if i < len(RELACOES) else None
        mand = MANDATORIO[i] if i < len(MANDATORIO) else None
        ws.append([dim, rel, mand])
    ws.sheet_state = "hidden"
    return ws


def _add_validations(ws, listas_ws) -> None:
    last_dim = len(DIMENSOES) + 1
    last_rel = len(RELACOES) + 1
    last_mand = len(MANDATORIO) + 1

    dv_dim = DataValidation(
        type="list",
        formula1=f"='_Listas'!$A$2:$A${last_dim}",
        allow_blank=True,
        showDropDown=False,
    )
    dv_dim.error = "Selecione uma dimensao da lista suspensa."
    dv_dim.errorTitle = "Dimensao invalida"
    dv_dim.prompt = "Escolha a dimensao do criterio."
    dv_dim.promptTitle = "Dimensao"

    dv_rel = DataValidation(
        type="list",
        formula1=f"='_Listas'!$B$2:$B${last_rel}",
        allow_blank=True,
        showDropDown=False,
    )
    dv_rel.error = "Use ↑ Positiva, ↓ Negativa ou ↕ Condicional."
    dv_rel.errorTitle = "Relacao invalida"
    dv_rel.prompt = "Indique se maior valor aumenta ou reduz prioridade."
    dv_rel.promptTitle = "Relacao"

    dv_mand = DataValidation(
        type="list",
        formula1=f"='_Listas'!$C$2:$C${last_mand}",
        allow_blank=True,
        showDropDown=False,
    )
    dv_mand.error = "Selecione Sim ou Nao."
    dv_mand.errorTitle = "Mandatorio invalido"
    dv_mand.prompt = "Criterio obrigatorio na rodada?"
    dv_mand.promptTitle = "Mandatorio"

    for dv in (dv_dim, dv_rel, dv_mand):
        ws.add_data_validation(dv)

    dv_dim.add(f"A2:A{MAX_DATA_ROW}")
    dv_rel.add(f"D2:D{MAX_DATA_ROW}")
    dv_mand.add(f"G2:G{MAX_DATA_ROW}")


def _aba_matriz_preenchimento(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Tabela de Premissas e Critérios"
    ws.append(HEADERS)
    _style_header(ws, len(HEADERS))
    _col_widths(ws)
    _hide_cols_from_h(ws)
    _apply_grid(ws, 2, MAX_DATA_ROW)
    _add_validations(ws, wb["_Listas"])


def _format_example_row(ws, row: int) -> None:
    ws.cell(row=row, column=1).font = DIM_FONT
    ws.cell(row=row, column=4).alignment = CENTER
    ws.cell(row=row, column=7).alignment = CENTER
    mand = ws.cell(row=row, column=7)
    if mand.value == "Sim":
        mand.fill = MAND_FILL
        mand.font = MAND_FONT


def _aba_exemplo(wb: Workbook) -> None:
    ws = wb.create_sheet("Exemplo de preenchimento")

    instrucoes = [
        "INSTRUCOES DE PREENCHIMENTO — Tabela de Premissas e Criterios (Analise AHP)",
        "",
        "1. Use a aba «Tabela de Premissas e Critérios» para cadastrar a etapa conceitual da sua analise.",
        "2. Preencha UMA LINHA por criterio. A premissa e obrigatoria: explica por que o criterio importa (logica da decisao).",
        "3. Dimensao, Relacao e Mandatorio: selecione sempre na lista suspensa (nao digite texto livre).",
        "4. Relacao: ↑ Positiva = quanto maior o indicador, maior a prioridade; ↓ Negativa = quanto pior o indicador, maior a prioridade; ↕ Condicional = depende do contexto da premissa.",
        "5. Metricas: descreva como medir (unidade, escala, fonte de dado). Fonte: referencia bibliografica, norma ou documento.",
        "6. Mandatorio = Sim apenas para criterios que DEVEM entrar na rodada de hierarquizacao.",
        "7. Apos preencher, salve o arquivo e faca upload no modulo AHP do SLT (etapa conceitual).",
        "8. Nao altere os titulos da linha 1 nem oculte as colunas A–G.",
        "",
        "Exemplos reais (tabela de referencia PLI-SP) — abaixo:",
    ]

    row = 1
    for line in instrucoes:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        cell = ws.cell(row=row, column=1, value=line)
        cell.alignment = WRAP_TOP
        if row == 1:
            cell.font = Font(bold=True, size=12, color="1F4E78")
        else:
            cell.font = INSTR_FONT
        if row <= 10:
            cell.fill = INSTR_FILL
        ws.row_dimensions[row].height = 22 if line else 10
        row += 1

    row += 1
    header_row = row
    ws.append(HEADERS)
    _style_header(ws, len(HEADERS))

    # Subconjunto representativo da matriz PLI-SP (dados reais)
    exemplos = CRITERIOS[:12]  # tecnica + inicio financeiro/economico
    for c in exemplos:
        rel_label = REL.get(c[3], c[3])
        ws.append([c[0], c[1], c[2], rel_label, c[4], c[5], c[6]])
        _format_example_row(ws, ws.max_row)

    _col_widths(ws)
    _hide_cols_from_h(ws)
    _apply_grid(ws, header_row + 1, ws.max_row)
    ws.freeze_panes = f"A{header_row + 1}"


def _export_csv_template() -> None:
    import csv

    OUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    with OUT_CSV.open("w", encoding="utf-8-sig", newline="") as handle:
        csv.writer(handle).writerow(HEADERS)


def main() -> None:
    wb = Workbook()
    _aba_listas(wb)
    _aba_matriz_preenchimento(wb)
    _aba_exemplo(wb)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(OUT)
    except PermissionError:
        stamp = __import__("datetime").datetime.now().strftime("%Y%m%d_%H%M%S")
        alt = OUT.with_name(f"Modelo_Matriz_Criterios_Premissas_Analise_{stamp}.xlsx")
        wb.save(alt)
        print(f"Aviso: arquivo padrao bloqueado; salvo em {alt}")
        return

    _export_csv_template()

    print(f"Modelo gerado: {OUT}")
    print(f"Modelo CSV: {OUT_CSV}")
    print(f"  - Abas: Tabela de Premissas e Critérios, Exemplo de preenchimento, _Listas (oculta)")
    print(f"  - Validacao: Dimensao ({len(DIMENSOES)}), Relacao ({len(RELACOES)}), Mandatorio (2)")
    print(f"  - Linhas editaveis na matriz: 2–{MAX_DATA_ROW}")


if __name__ == "__main__":
    main()
