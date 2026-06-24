"""Gera Catalogo_Hierarquico_SLT.xlsx — catálogo operacional (revisar internamente se necessário)."""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "data" / "Catalogo_Hierarquico_SLT.xlsx"

HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14)
WRAP = Alignment(wrap_text=True, vertical="top")


def style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_sheet(wb, title, headers, rows, widths):
    ws = wb.create_sheet(title)
    ws.append(headers)
    style_header(ws)
    for row in rows:
        ws.append(row)
    set_widths(ws, widths)
    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).alignment = WRAP
    ws.freeze_panes = "A2"
    return ws


wb = Workbook()
ws0 = wb.active
ws0.title = "Instrucoes"
ws0["A1"] = "Catálogo Hierárquico SLT — versão operacional"
ws0["A1"].font = TITLE_FONT
lines = [
    "",
    "Catálogo de trabalho do sistema. Nomes definidos para implementação; revisar internamente quando necessário.",
    "",
    "Fluxo de cadastro (demandante):",
    "  1. Diretoria SLT",
    "  2. Plano",
    "  3. Projeto (nome, geometria obrigatória, contato)",
    "  4. Se PLI → Frente de atuação | Se PEF → Eixo ferroviário",
    "  5. Complementos: Modal, Tipologia, Carteira (opcional)",
    "",
    "Referências: Decreto 69.376/2025; pli.semil.sp.gov.br/frentes-de-atucao/; PEF-SP / SP nos Trilhos.",
]
for i, line in enumerate(lines, 2):
    ws0.cell(i, 1, line)
ws0.column_dimensions["A"].width = 90

add_sheet(
    wb,
    "Diretorias",
    ["id", "nome_oficial", "descricao", "decreto_ref", "ativo"],
    [
        (
            "DIR-PLAN",
            "Diretoria de Planejamento de Logística e Transportes",
            "Planos, intermodalidade, estudos e financiamento de empreendimentos logísticos",
            "Art. 36",
            "SIM",
        ),
        (
            "DIR-GEST",
            "Diretoria de Gestão de Transportes",
            "Operação, manutenção e conservação; hidrovia; travessias; convênios de delegação",
            "Art. 37",
            "SIM",
        ),
        (
            "DIR-INFRA",
            "Diretoria de Infraestrutura de Transportes",
            "Obras, carteira de projetos, contratações, sistema georreferenciado de transportes",
            "Art. 38",
            "SIM",
        ),
    ],
    [12, 44, 52, 12, 8],
)

add_sheet(
    wb,
    "Planos",
    [
        "id",
        "nome_oficial",
        "sigla",
        "diretoria_id",
        "horizonte",
        "descricao",
        "classificacao_pos_projeto",
        "ativo",
    ],
    [
        (
            "PLANO-PLI",
            "Plano de Logística e Investimentos do Estado de São Paulo",
            "PLI-SP 2050",
            "DIR-PLAN",
            "2050",
            "Plano estadual multimodal de logística e investimentos em transportes",
            "Frente de atuação",
            "SIM",
        ),
        (
            "PLANO-PEF",
            "Plano Estratégico Ferroviário do Estado de São Paulo",
            "PEF-SP 2050",
            "DIR-PLAN",
            "2050",
            "Plano setorial ferroviário articulado ao PLI-SP; passageiros, carga e malha ociosa",
            "Eixo ferroviário",
            "SIM",
        ),
    ],
    [12, 48, 14, 14, 10, 48, 22, 8],
)

add_sheet(
    wb,
    "Frentes_Atuacao_PLI",
    ["id", "nome_oficial", "descricao", "plano_id", "ativo"],
    [
        (
            "FRENTE-01",
            "Polos de Desenvolvimento Regional",
            "Identificação e estruturação de áreas estratégicas para investimentos, empregos e dinamização regional.",
            "PLANO-PLI",
            "SIM",
        ),
        (
            "FRENTE-02",
            "Ligação Planalto-Litoral",
            "Conexão interior–litoral; escoamento de cargas e acesso aos portos.",
            "PLANO-PLI",
            "SIM",
        ),
        (
            "FRENTE-03",
            "Ramais Ferroviários",
            "Mapeamento, desenvolvimento e reativação de trechos ferroviários com potencial logístico.",
            "PLANO-PLI",
            "SIM",
        ),
        (
            "FRENTE-04",
            "Transporte para Biocombustíveis",
            "Logística e infraestrutura associadas a biocombustíveis e dutovias.",
            "PLANO-PLI",
            "SIM",
        ),
        (
            "FRENTE-05",
            "Ampliação de Hidrovias",
            "Expansão e modernização da navegação fluvial e da logística hidroviária.",
            "PLANO-PLI",
            "SIM",
        ),
        (
            "FRENTE-06",
            "Mobilidade Paulista",
            "Mobilidade e integração territorial com inclusão social e sustentabilidade.",
            "PLANO-PLI",
            "SIM",
        ),
    ],
    [14, 34, 55, 12, 8],
)

# Eixos PEF — estrutura proposta alinhada a PEF-SP, PLI (TIC/Expresso Carga) e SP nos Trilhos
add_sheet(
    wb,
    "Eixos_PEF",
    ["id", "nome_oficial", "descricao", "plano_id", "ativo"],
    [
        (
            "EIXO-PEF-01",
            "Reativação de Malha Ociosa e Ramais",
            "Recuperação e reativação de trechos ferroviários abandonados ou subutilizados no Estado.",
            "PLANO-PEF",
            "SIM",
        ),
        (
            "EIXO-PEF-02",
            "Trens Intercidades",
            "Corredores de passageiros inter-regionais sobre trilhos (TIC), integrando polos metropolitanos e interior.",
            "PLANO-PEF",
            "SIM",
        ),
        (
            "EIXO-PEF-03",
            "Carga Ferroviária e Expresso Carga",
            "Transporte de cargas por ferrovia, integração caminhão–trem e terminais de transferência.",
            "PLANO-PEF",
            "SIM",
        ),
        (
            "EIXO-PEF-04",
            "Anel Metropolitano Ferroviário",
            "Integração ferroviária metropolitana e corredores de ligação radial–metropolitana.",
            "PLANO-PEF",
            "SIM",
        ),
        (
            "EIXO-PEF-05",
            "Terminais e Interfaces Intermodais",
            "Terminais ferroviários, intercambio modal e plataformas logísticas sobre trilhos.",
            "PLANO-PEF",
            "SIM",
        ),
        (
            "EIXO-PEF-06",
            "Shortlines e Ferrovias Regionais",
            "Linhas de trajeto curto, autorizações ferroviárias e operações regionais (marco PL ferrovias SP).",
            "PLANO-PEF",
            "SIM",
        ),
        (
            "EIXO-PEF-07",
            "Novos Corredores Ferroviários",
            "Expansão greenfield e novos traçados ferroviários de longo prazo.",
            "PLANO-PEF",
            "SIM",
        ),
    ],
    [14, 36, 58, 12, 8],
)

# Corredores TIC (subclassificação opcional sob EIXO-PEF-02)
add_sheet(
    wb,
    "Corredores_TIC",
    ["id", "nome_oficial", "eixo_pef_id", "descricao", "ativo"],
    [
        ("TIC-NORTE", "Trem Intercidades São Paulo–Campinas (Eixo Norte)", "EIXO-PEF-02", "Ligação capital–Campinas e eixo norte", "SIM"),
        ("TIC-LESTE", "Trem Intercidades São Paulo–Sorocaba (Eixo Leste)", "EIXO-PEF-02", "Ligação capital–Sorocaba", "SIM"),
        ("TIC-SUL", "Trem Intercidades São Paulo–Santos (Eixo Sul)", "EIXO-PEF-02", "Ligação capital–litoral / porto", "SIM"),
        ("TIC-OESTE", "Trem Intercidades São Paulo–São José dos Campos (Eixo Oeste)", "EIXO-PEF-02", "Ligação capital–Vale do Paraíba", "SIM"),
        ("TIC-MAR-SOR", "Trem Intercidades Marília–Sorocaba", "EIXO-PEF-02", "Corredor interior oeste", "SIM"),
        ("TIC-SOR-CAM", "Trem Intercidades Sorocaba–Campinas", "EIXO-PEF-02", "Corredor interior central", "SIM"),
        ("TIC-CAM-RIB", "Trem Intercidades Campinas–Ribeirão Preto", "EIXO-PEF-02", "Corredor noroeste paulista", "SIM"),
        ("TIC-RIB-FRA", "Trem Intercidades Ribeirão Preto–Franca", "EIXO-PEF-02", "Extensão noroeste", "SIM"),
        ("TIC-SAN-CAJ", "Trem Intercidades Santos–Cajati", "EIXO-PEF-02", "Ligação litoral–serra", "SIM"),
        ("TIC-CAM-ARA", "Trem Intercidades Campinas–Araraquara", "EIXO-PEF-02", "Corredor leste interior", "SIM"),
        ("TIC-SJC-TAU", "Trem Intercidades São José dos Campos–Taubaté", "EIXO-PEF-02", "Corredor Vale do Paraíba", "SIM"),
    ],
    [14, 42, 14, 40, 8],
)

add_sheet(
    wb,
    "Modais",
    ["id", "nome", "descricao", "ativo"],
    [
        ("MOD-RODO", "Rodoviário", "Rodovias, acessos, contornos, duplicações", "SIM"),
        ("MOD-FERR", "Ferroviário", "Malha ferroviária, terminais, ramais", "SIM"),
        ("MOD-HIDR", "Hidroviário", "Hidrovias e terminais fluviais", "SIM"),
        ("MOD-PORT", "Portuário", "Portos e acessos portuários", "SIM"),
        ("MOD-AERO", "Aeroportuário", "Aeroportos e integração logística aérea", "SIM"),
        ("MOD-INTER", "Intermodal", "Terminais e elos multimodais", "SIM"),
    ],
    [12, 18, 48, 8],
)

add_sheet(
    wb,
    "Tipologias_Intervencao",
    ["id", "nome", "descricao", "ativo"],
    [
        ("TIP-OBRA", "Obra de infraestrutura", "Construção ou ampliação física", "SIM"),
        ("TIP-DUP", "Duplicação / melhoria viária", "Duplicação, faixas, contornos rodoviários", "SIM"),
        ("TIP-REAT", "Reativação ferroviária", "Recuperação e retomada de operação sobre trilhos", "SIM"),
        ("TIP-TERM", "Terminal intermodal", "Terminal de cargas, passageiros ou transferência modal", "SIM"),
        ("TIP-HIDR", "Intervenção hidroviária", "Dragagem, eclusas, estruturas de apoio à navegação", "SIM"),
        ("TIP-PORT", "Intervenção portuária", "Berços, acessos, infraestrutura portuária", "SIM"),
        ("TIP-EST", "Estudo e projeto", "Estudo de viabilidade, plano básico, projeto executivo", "SIM"),
        ("TIP-CONS", "Conservação e operação", "Manutenção, conservação especial, operação", "SIM"),
        ("TIP-DIG", "Plataforma e dados", "Sistemas, automação, geoinformação", "SIM"),
    ],
    [12, 28, 48, 8],
)

add_sheet(
    wb,
    "Carteiras_Projetos",
    ["id", "nome", "plano_id", "referencia", "ativo"],
    [
        ("CART-PLI-2026", "Carteira PLI-SP 2026", "PLANO-PLI", "Ciclo corrente de priorização PLI", "SIM"),
        ("CART-PEF-2026", "Carteira PEF-SP 2026", "PLANO-PEF", "Ciclo corrente de priorização ferroviária", "SIM"),
    ],
    [16, 32, 12, 40, 8],
)

add_sheet(
    wb,
    "Tipos_Grupo",
    ["tipo_grupo_id", "rotulo_ui", "descricao", "obrigatorio", "aba_lista", "ativo"],
    [
        ("GRUPO-FRENTE", "Frente de atuação", "Classificação PLI após cadastro do Projeto", "SIM se PLI", "Frentes_Atuacao_PLI", "SIM"),
        ("GRUPO-EIXO", "Eixo ferroviário", "Classificação PEF após cadastro do Projeto", "SIM se PEF", "Eixos_PEF", "SIM"),
        ("GRUPO-TIC", "Corredor TIC", "Subclassificação opcional para projetos TIC", "NAO", "Corredores_TIC", "SIM"),
        ("GRUPO-MODAL", "Modal", "Recorte multimodal", "SUGERIDO", "Modais", "SIM"),
        ("GRUPO-TIPOL", "Tipologia de intervenção", "Natureza da intervenção", "NAO", "Tipologias_Intervencao", "SIM"),
        ("GRUPO-CART", "Carteira de projetos", "Vínculo à carteira de investimentos", "NAO", "Carteiras_Projetos", "SIM"),
    ],
    [16, 26, 42, 14, 24, 8],
)

add_sheet(
    wb,
    "Fluxo_Cadastro",
    ["passo", "rotulo_ui", "tipo_interno", "obrigatorio", "fonte_aba", "observacoes"],
    [
        (1, "Diretoria", "diretoria", "SIM", "Diretorias", ""),
        (2, "Plano", "plano", "SIM", "Planos", ""),
        (3, "Projeto", "projeto", "SIM", "—", "Nome, geometria, contato"),
        (4, "Frente de atuação", "frente", "SIM se PLI", "Frentes_Atuacao_PLI", "Rótulo dinâmico: ver coluna classificacao_pos_projeto em Planos"),
        (4, "Eixo ferroviário", "eixo_pef", "SIM se PEF", "Eixos_PEF", "Mutuamente exclusivo com frente"),
        (5, "Corredor TIC", "corredor_tic", "NAO", "Corredores_TIC", "Somente se eixo = Trens Intercidades"),
        (6, "Modal", "modal", "SUGERIDO", "Modais", ""),
        (7, "Tipologia", "tipologia", "NAO", "Tipologias_Intervencao", ""),
        (8, "Carteira de projetos", "carteira", "NAO", "Carteiras_Projetos", ""),
    ],
    [8, 28, 18, 14, 26, 42],
)

add_sheet(
    wb,
    "Entidades_Demandantes",
    ["id", "nome", "tipo", "ativo"],
    [
        ("DEM-DER", "Departamento de Estradas de Rodagem (DER-SP)", "Estadual", "SIM"),
        ("DEM-CPTM", "CPTM", "Estadual", "SIM"),
        ("DEM-CDSS", "Companhia Docas de São Sebastião", "Estadual", "SIM"),
        ("DEM-MUN", "Município / consórcio intermunicipal", "Municipal", "SIM"),
        ("DEM-PRIV", "Setor privado / concessionária", "Privado", "SIM"),
        ("DEM-OUT-SEC", "Outra secretaria estadual", "Estadual", "SIM"),
        ("DEM-OUT", "Outra entidade", "Outro", "SIM"),
    ],
    [12, 42, 14, 8],
)

wb.save(OUT)
print(f"Gerado: {OUT}")
