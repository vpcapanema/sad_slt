from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


OUTPUT_PATH = Path(
    "data/ESTRUTURA_MODELO_HIERARQUIZACAO_PLI.xlsx"
)
SOURCE_PATH = Path(
    "data/Matriz_Criterios_Premissas_PLI-SP_20260608_195036.xlsx"
)


def make_row(*cells):
    return list(cells)


COMMON_HEADERS = [
    "Código",
    "Etapa",
    "Subetapa",
    "Dimensão",
    "Critério",
    "Premissa",
    "Variável",
    "Relação com o escopo",
    "Dado-fonte",
    "Dado derivado",
    "Unidade de medida / métrica",
    "Fonte",
    "Observações",
]

SYNTHESIS_HEADERS = [
    "Tipo de registro",
    "Etapa",
    "Código de origem",
    "Subetapa",
    "Dimensão",
    "Critério",
    "Função no fluxo",
    "Quantidade",
    "Classificação de origem",
    "Mandatório",
    "Observações",
]

RELATION_MAP = {
    "↑ Positiva": "positiva",
    "↓ Negativa": "negativa",
    "↕ Condicional": "condicional",
}

ETAPA_1_ITEM_CONFIG = [
    {
        "kind": "section",
        "code": "1",
        "subetapa": None,
        "dimensao": None,
        "obs": "Triagem inicial do projeto.",
    },
    {
        "kind": "section",
        "code": "1.1",
        "subetapa": "Restrição",
        "dimensao": None,
        "obs": "Critérios impeditivos: se atendidos, o projeto é barrado.",
    },
    {
        "kind": "section",
        "code": "1.1.1",
        "subetapa": "Restrição",
        "dimensao": "Ambiental",
        "obs": "Agrupa critérios impeditivos ambientais e locacionais.",
    },
    {
        "kind": "criterion",
        "code": "1.1.1.1",
        "subetapa": "Restrição",
        "dimensao": "Ambiental",
        "source_key": (
            "Menor conflito locacional com áreas sensíveis ou protegidas"
        ),
        "criterio": "Sobreposição com áreas protegidas",
        "premissa": (
            "Projetos que sobrepõem áreas protegidas possuem licenciamento "
            "ambiental complexo, o que desestimula sua priorização e pode "
            "representar desafios indesejados do ponto de vista de execução."
        ),
        "variavel": "Sobreposição com áreas protegidas",
        "relacao": "negativa",
        "dado_fonte": (
            "Shapefile com o perímetro das áreas protegidas do Estado de "
            "São Paulo."
        ),
        "dado_derivado": (
            "Presença ou ausência de sobreposição entre o projeto e o "
            "perímetro das áreas protegidas."
        ),
        "unidade": "presença/ausência",
        "fonte": (
            "A definir: base oficial de áreas protegidas do Estado de São "
            "Paulo."
        ),
    },
    {
        "kind": "criterion",
        "code": "1.1.1.2",
        "subetapa": "Restrição",
        "dimensao": "Ambiental",
        "criterio": "Menor complexidade licenciatória para implantação",
    },
    {
        "kind": "section",
        "code": "1.1.2",
        "subetapa": "Restrição",
        "dimensao": "Jurídico-institucional",
        "obs": "Agrupa critérios impeditivos de conformidade jurídica.",
    },
    {
        "kind": "criterion",
        "code": "1.1.2.1",
        "subetapa": "Restrição",
        "dimensao": "Jurídico-institucional",
        "criterio": "Menor exposição a entraves jurídicos e jurisdicionais",
    },
    {
        "kind": "section",
        "code": "1.1.3",
        "subetapa": "Restrição",
        "dimensao": "Territorial",
        "obs": "Agrupa critérios impeditivos de aderência territorial.",
    },
    {
        "kind": "criterion",
        "code": "1.1.3.1",
        "subetapa": "Restrição",
        "dimensao": "Territorial",
        "criterio": (
            "Maior compatibilidade territorial e urbanística com o "
            "planejamento local"
        ),
    },
    {
        "kind": "section",
        "code": "1.2",
        "subetapa": "Risco",
        "dimensao": None,
        "obs": "Critérios não impeditivos: o projeto segue com ressalvas.",
    },
    {
        "kind": "section",
        "code": "1.2.1",
        "subetapa": "Risco",
        "dimensao": "Climático-ambiental",
        "obs": "Agrupa riscos ambientais, climáticos e socioambientais.",
    },
    {
        "kind": "criterion",
        "code": "1.2.1.1",
        "subetapa": "Risco",
        "dimensao": "Climático-ambiental",
        "source_key": (
            "Maior contribuição para resiliência da rede a eventos extremos"
        ),
        "criterio": (
            "Sobreposição com zonas de amortecimento de áreas protegidas"
        ),
        "premissa": (
            "Projetos que se sobrepõem às zonas de amortecimento de áreas "
            "protegidas podem ter processo de licenciamento, execução ou "
            "operação mais complexos, o que representa um desafio "
            "desfavorável do ponto de vista de execução."
        ),
        "variavel": (
            "Sobreposição com zonas de amortecimento de áreas protegidas"
        ),
        "relacao": "negativa",
        "dado_fonte": (
            "Shapefile com o perímetro das áreas protegidas do Estado de "
            "São Paulo."
        ),
        "dado_derivado": (
            "Zonas de amortecimento obtidas por buffer de 10 km a partir do "
            "perímetro das áreas protegidas e presença ou ausência de "
            "sobreposição com o projeto."
        ),
        "unidade": "presença/ausência",
        "fonte": (
            "A definir: base oficial de áreas protegidas do Estado de São "
            "Paulo."
        ),
    },
    {
        "kind": "criterion",
        "code": "1.2.1.2",
        "subetapa": "Risco",
        "dimensao": "Climático-ambiental",
        "criterio": (
            "Menor conflito socioambiental com comunidades tradicionais"
        ),
    },
    {
        "kind": "section",
        "code": "1.2.2",
        "subetapa": "Risco",
        "dimensao": "Demanda",
        "obs": "Agrupa riscos de incerteza do benefício esperado.",
    },
    {
        "kind": "criterion",
        "code": "1.2.2.1",
        "subetapa": "Risco",
        "dimensao": "Demanda",
        "criterio": "Menor incerteza quanto à demanda futura do projeto",
    },
    {
        "kind": "section",
        "code": "1.2.3",
        "subetapa": "Risco",
        "dimensao": "Execução",
        "obs": "Agrupa riscos de implementação e dependências externas.",
    },
    {
        "kind": "criterion",
        "code": "1.2.3.1",
        "subetapa": "Risco",
        "dimensao": "Execução",
        "criterio": "Menor risco de atraso e sobrecusto na implantação",
    },
    {
        "kind": "criterion",
        "code": "1.2.3.2",
        "subetapa": "Risco",
        "dimensao": "Execução",
        "criterio": (
            "Menor carga de desapropriações e interferências físicas"
        ),
    },
    {
        "kind": "criterion",
        "code": "1.2.3.3",
        "subetapa": "Risco",
        "dimensao": "Execução",
        "criterio": (
            "Menor dependência de projetos predecessores ou de entregas "
            "externas"
        ),
    },
]

ETAPA_2_ITEM_CONFIG = [
    {
        "kind": "section",
        "code": "2",
        "dimensao": None,
        "obs": "Hierarquização pareada por análise multicritério.",
    },
    {
        "kind": "section",
        "code": "2.1",
        "dimensao": "Técnica",
        "obs": "Agrupa critérios de desempenho, operação e condição da rede.",
    },
    {
        "kind": "criterion",
        "code": "2.1.1",
        "dimensao": "Técnica",
        "criterio": "Proximidade com segmentos rodoviários de VDM alto",
    },
    {
        "kind": "criterion",
        "code": "2.1.2",
        "dimensao": "Técnica",
        "criterio": "Proximidade com segmentos de saturação elevada",
    },
    {
        "kind": "criterion",
        "code": "2.1.3",
        "dimensao": "Técnica",
        "criterio": "Proximidade com segmentos de lentidão recorrente",
    },
    {
        "kind": "criterion",
        "code": "2.1.4",
        "dimensao": "Técnica",
        "criterio": "Proximidade com segmentos de pavimento degradado",
    },
    {
        "kind": "criterion",
        "code": "2.1.5",
        "dimensao": "Técnica",
        "criterio": "Proximidade com segmentos de geometria deficiente",
    },
    {
        "kind": "criterion",
        "code": "2.1.6",
        "dimensao": "Técnica",
        "criterio": "Proximidade com segmentos de forte sobrecarga sazonal",
    },
    {
        "kind": "section",
        "code": "2.2",
        "dimensao": "Econômico-financeira",
        "obs": (
            "Agrupa critérios de retorno, custo, competitividade e "
            "eficiência econômica."
        ),
    },
    {
        "kind": "criterion",
        "code": "2.2.1",
        "dimensao": "Econômico-financeira",
        "criterio": "Menor custo de investimento por benefício esperado",
    },
    {
        "kind": "criterion",
        "code": "2.2.2",
        "dimensao": "Econômico-financeira",
        "criterio": "Menor custo operacional ao longo da vida útil",
    },
    {
        "kind": "criterion",
        "code": "2.2.3",
        "dimensao": "Econômico-financeira",
        "criterio": "Maior retorno econômico por unidade investida",
    },
    {
        "kind": "criterion",
        "code": "2.2.4",
        "dimensao": "Econômico-financeira",
        "criterio": (
            "Maior vantagem locacional frente a corredores de menor custo "
            "logístico"
        ),
    },
    {
        "kind": "criterion",
        "code": "2.2.5",
        "dimensao": "Econômico-financeira",
        "criterio": "Maior benefício social líquido do empreendimento",
    },
    {
        "kind": "criterion",
        "code": "2.2.6",
        "dimensao": "Econômico-financeira",
        "criterio": "Localização em áreas de alta massa econômica",
    },
    {
        "kind": "criterion",
        "code": "2.2.7",
        "dimensao": "Econômico-financeira",
        "criterio": "Maior ganho de competitividade para a produção atendida",
    },
    {
        "kind": "criterion",
        "code": "2.2.8",
        "dimensao": "Econômico-financeira",
        "criterio": "Maior acessibilidade temporal aos destinos relevantes",
    },
    {
        "kind": "criterion",
        "code": "2.2.9",
        "dimensao": "Econômico-financeira",
        "criterio": "Maior potencial de indução econômica regional",
    },
    {
        "kind": "criterion",
        "code": "2.2.10",
        "dimensao": "Econômico-financeira",
        "criterio": (
            "Maior captura de cargas hoje sem alternativa logística eficiente"
        ),
    },
    {
        "kind": "criterion",
        "code": "2.2.11",
        "dimensao": "Econômico-financeira",
        "criterio": (
            "Maior suporte locacional e funcional a cadeias estratégicas"
        ),
    },
    {
        "kind": "criterion",
        "code": "2.2.12",
        "dimensao": "Econômico-financeira",
        "criterio": (
            "Maior acessibilidade funcional a eixos hidroviários eficientes"
        ),
    },
    {
        "kind": "criterion",
        "code": "2.2.13",
        "dimensao": "Econômico-financeira",
        "criterio": (
            "Maior acessibilidade funcional à malha ferroviária "
            "estratégica"
        ),
    },
    {
        "kind": "section",
        "code": "2.3",
        "dimensao": "Social",
        "obs": (
            "Agrupa critérios de inclusão, acesso e alcance social do "
            "projeto."
        ),
    },
    {
        "kind": "criterion",
        "code": "2.3.1",
        "dimensao": "Social",
        "criterio": (
            "Localização em áreas de maior vulnerabilidade territorial"
        ),
    },
    {
        "kind": "criterion",
        "code": "2.3.2",
        "dimensao": "Social",
        "criterio": "Maior população efetivamente alcançada pelo projeto",
    },
    {
        "kind": "criterion",
        "code": "2.3.3",
        "dimensao": "Social",
        "criterio": "Maior melhoria de acesso para populações subatendidas",
    },
    {
        "kind": "criterion",
        "code": "2.3.4",
        "dimensao": "Social",
        "criterio": "Maior melhoria de acesso a saúde e educação",
    },
    {
        "kind": "criterion",
        "code": "2.3.5",
        "dimensao": "Social",
        "criterio": (
            "Maior acessibilidade funcional a polos logísticos estratégicos"
        ),
    },
    {
        "kind": "criterion",
        "code": "2.3.6",
        "dimensao": "Social",
        "criterio": (
            "Maior conexão de comunidades isoladas a oportunidades e "
            "serviços"
        ),
    },
    {
        "kind": "section",
        "code": "2.4",
        "dimensao": "Segurança",
        "obs": (
            "Agrupa critérios de redução de acidentes e de exposição a "
            "sinistros graves."
        ),
    },
    {
        "kind": "criterion",
        "code": "2.4.1",
        "dimensao": "Segurança",
        "criterio": "Proximidade com segmentos de alta gravidade de acidentes",
    },
    {
        "kind": "criterion",
        "code": "2.4.2",
        "dimensao": "Segurança",
        "criterio": (
            "Proximidade com segmentos de alta incidência de acidentes com "
            "usuários vulneráveis"
        ),
    },
    {
        "kind": "criterion",
        "code": "2.4.3",
        "dimensao": "Segurança",
        "criterio": (
            "Maior necessidade de mitigação em eixos de cargas perigosas"
        ),
    },
    {
        "kind": "criterion",
        "code": "2.4.4",
        "dimensao": "Segurança",
        "criterio": (
            "Proximidade com concentração elevada de pontos críticos de "
            "acidentes"
        ),
    },
    {
        "kind": "section",
        "code": "2.5",
        "dimensao": "Ambiental",
        "obs": "Agrupa critérios de desempenho ambiental e eficiência modal.",
    },
    {
        "kind": "criterion",
        "code": "2.5.1",
        "dimensao": "Ambiental",
        "criterio": (
            "Maior potencial de redução de emissões associadas à localização "
            "e ao efeito de rede"
        ),
    },
    {
        "kind": "criterion",
        "code": "2.5.2",
        "dimensao": "Ambiental",
        "criterio": (
            "Maior potencial de redução de poluição local em áreas expostas"
        ),
    },
    {
        "kind": "criterion",
        "code": "2.5.3",
        "dimensao": "Ambiental",
        "criterio": "Maior eficiência energética do sistema atendido",
    },
    {
        "kind": "criterion",
        "code": "2.5.4",
        "dimensao": "Ambiental",
        "criterio": "Maior potencial de migração para modais mais eficientes",
    },
    {
        "kind": "section",
        "code": "2.6",
        "dimensao": "Territorial",
        "obs": (
            "Agrupa critérios de conectividade, inserção espacial e "
            "integração logística."
        ),
    },
    {
        "kind": "criterion",
        "code": "2.6.1",
        "dimensao": "Territorial",
        "criterio": (
            "Proximidade com segmentos de alto conflito urbano-regional"
        ),
    },
    {
        "kind": "criterion",
        "code": "2.6.2",
        "dimensao": "Territorial",
        "criterio": (
            "Proximidade com segmentos de alta interferência "
            "urbano-portuária"
        ),
    },
    {
        "kind": "criterion",
        "code": "2.6.3",
        "dimensao": "Territorial",
        "criterio": (
            "Maior acessibilidade funcional a nós intermodais estratégicos"
        ),
    },
    {
        "kind": "criterion",
        "code": "2.6.4",
        "dimensao": "Territorial",
        "criterio": (
            "Maior capacidade de reduzir vazios logísticos e conectar "
            "regiões pouco integradas"
        ),
    },
    {
        "kind": "criterion",
        "code": "2.6.5",
        "dimensao": "Territorial",
        "criterio": (
            "Proximidade com polos geradores e atratores de tráfego "
            "relevantes"
        ),
    },
]

ETAPA_3_ITEM_CONFIG = [
    {
        "kind": "section",
        "code": "3",
        "dimensao": None,
        "obs": "Ajuste de prioridade após a hierarquização pareada.",
    },
    {
        "kind": "section",
        "code": "3.1",
        "dimensao": "Prontidão",
        "obs": (
            "Agrupa critérios de maturidade e velocidade de entrada em "
            "operação."
        ),
    },
    {
        "kind": "criterion",
        "code": "3.1.1",
        "dimensao": "Prontidão",
        "criterio": "Maior prontidão para implantação",
    },
    {
        "kind": "criterion",
        "code": "3.1.2",
        "dimensao": "Prontidão",
        "criterio": "Menor prazo até a entrada em operação",
    },
    {
        "kind": "section",
        "code": "3.2",
        "dimensao": "Viabilidade",
        "obs": (
            "Agrupa critérios de executabilidade técnica, institucional e "
            "financeira."
        ),
    },
    {
        "kind": "criterion",
        "code": "3.2.1",
        "dimensao": "Viabilidade",
        "criterio": (
            "Menor complexidade técnica e institucional do empreendimento"
        ),
    },
    {
        "kind": "criterion",
        "code": "3.2.2",
        "dimensao": "Viabilidade",
        "criterio": "Maior atratividade para financiamento privado",
    },
    {
        "kind": "criterion",
        "code": "3.2.3",
        "dimensao": "Viabilidade",
        "criterio": (
            "Maior consenso institucional para viabilização do projeto"
        ),
    },
    {
        "kind": "section",
        "code": "3.3",
        "dimensao": "Oportunidade",
        "obs": (
            "Agrupa critérios de alinhamento estratégico e janela "
            "política-social."
        ),
    },
    {
        "kind": "criterion",
        "code": "3.3.1",
        "dimensao": "Oportunidade",
        "criterio": "Maior aderência estratégica aos planos vigentes",
    },
    {
        "kind": "criterion",
        "code": "3.3.2",
        "dimensao": "Oportunidade",
        "criterio": (
            "Maior legitimidade social e participativa do empreendimento"
        ),
    },
]

FLOW_FUNCTION = {
    "Etapa 1": "Triagem",
    "Etapa 2": "Hierarquização",
    "Etapa 3": "Ajuste",
}

FLOW_OUTCOME = {
    "Etapa 1": "Classificar projeto como inapto, apto com ressalvas ou apto.",
    "Etapa 2": "Gerar posição relativa por mérito multicritério.",
    "Etapa 3": (
        "Refinar a posição relativa por prontidão, viabilidade e "
        "oportunidade."
    ),
}


def load_source_criteria() -> dict[str, dict[str, str | None]]:
    workbook = load_workbook(SOURCE_PATH, read_only=True, data_only=True)
    worksheet = workbook["Matriz Crit Premissas v2"]
    headers = [cell.value for cell in worksheet[1]]
    criteria = {}

    for values in worksheet.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(headers, values))
        criteria[row_data["Critério"]] = row_data

    return criteria


def build_criterion_row(
    item: dict[str, str | None],
    code: str,
    etapa: str,
    subetapa: str | None,
    dimensao: str,
    source_row: dict[str, str | None],
) -> list[str | None]:
    observacoes = (
        f"Classificação de origem: {source_row['Classificação']}; "
        f"Mandatório: {source_row['Mandatório']}"
    )
    criterio = item.get("criterio", source_row["Critério"])
    premissa = item.get("premissa", source_row["Premissa"])
    variavel = item.get("variavel", source_row["Variável"])
    relacao = item.get("relacao", RELATION_MAP[source_row["Relação"]])
    dado_fonte = item.get("dado_fonte", source_row["Dado"])
    dado_derivado = item.get("dado_derivado", source_row["Métricas"])
    unidade = item.get("unidade", source_row["Unidade de medida"])
    fonte = item.get("fonte", source_row["Fonte"])

    return make_row(
        code,
        etapa,
        subetapa,
        dimensao,
        criterio,
        premissa,
        variavel,
        relacao,
        dado_fonte,
        dado_derivado,
        unidade,
        fonte,
        observacoes,
    )


def build_section_row(
    code: str,
    etapa: str,
    subetapa: str | None,
    dimensao: str | None,
    observacao: str,
) -> list[str | None]:
    return make_row(
        code,
        etapa,
        subetapa,
        dimensao,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        observacao,
    )


def build_etapa_1_rows(
    source_criteria: dict[str, dict[str, str | None]],
) -> list[list[str | None]]:
    rows = []

    for item in ETAPA_1_ITEM_CONFIG:
        if item["kind"] == "section":
            rows.append(
                build_section_row(
                    item["code"],
                    "Etapa 1",
                    item["subetapa"],
                    item["dimensao"],
                    item["obs"],
                )
            )
            continue

        criterio = item.get("source_key", item["criterio"])
        rows.append(
            build_criterion_row(
                item,
                item["code"],
                "Etapa 1",
                item["subetapa"],
                item["dimensao"],
                source_criteria[criterio],
            )
        )

    return rows


def build_criterion_records(
    etapa: str,
    item_config: list[dict[str, str | None]],
    source_criteria: dict[str, dict[str, str | None]],
    default_subetapa: str | None = None,
) -> list[dict[str, str | None]]:
    records = []

    for item in item_config:
        if item["kind"] != "criterion":
            continue

        source_row = source_criteria[item.get("source_key", item["criterio"])]
        records.append(
            {
                "etapa": etapa,
                "codigo": item["code"],
                "subetapa": item.get("subetapa", default_subetapa),
                "dimensao": item["dimensao"],
                "criterio": item.get("criterio", source_row["Critério"]),
                "classificacao": source_row["Classificação"],
                "mandatorio": source_row["Mandatório"],
            }
        )

    return records


def build_synthesis_rows(
    records: list[dict[str, str | None]],
) -> list[list[str | None]]:
    rows = []
    etapas = ["Etapa 1", "Etapa 2", "Etapa 3"]

    for etapa in etapas:
        etapa_records = [
            record for record in records if record["etapa"] == etapa
        ]
        rows.append(
            make_row(
                "contagem",
                etapa,
                None,
                None,
                None,
                None,
                FLOW_FUNCTION[etapa],
                len(etapa_records),
                None,
                None,
                FLOW_OUTCOME[etapa],
            )
        )

        dimensoes = []
        for record in etapa_records:
            if record["dimensao"] not in dimensoes:
                dimensoes.append(record["dimensao"])

        for dimensao in dimensoes:
            dimension_records = [
                record for record in etapa_records
                if record["dimensao"] == dimensao
            ]
            rows.append(
                make_row(
                    "contagem-dimensão",
                    etapa,
                    None,
                    dimension_records[0]["subetapa"],
                    dimensao,
                    None,
                    FLOW_FUNCTION[etapa],
                    len(dimension_records),
                    None,
                    None,
                    "Quantidade de critérios alocados nesta dimensão.",
                )
            )

    rows.append(
        make_row(
            "marcador",
            "Todas",
            None,
            None,
            None,
            None,
            "Rastreabilidade",
            len(records),
            None,
            None,
            "Lista completa dos critérios com sua alocação no fluxo.",
        )
    )

    for record in records:
        rows.append(
            make_row(
                "critério",
                record["etapa"],
                record["codigo"],
                record["subetapa"],
                record["dimensao"],
                record["criterio"],
                FLOW_FUNCTION[record["etapa"]],
                None,
                record["classificacao"],
                record["mandatorio"],
                "Rastreado automaticamente a partir da matriz consolidada.",
            )
        )

    rows.append(
        make_row(
            "contagem-final",
            "Todas",
            None,
            None,
            None,
            None,
            "Cobertura",
            len(records),
            None,
            None,
            "Total de critérios alocados nas três etapas.",
        )
    )

    return rows


def save_workbook(workbook: Workbook, output_path: Path) -> Path:
    try:
        workbook.save(output_path)
        return output_path
    except PermissionError:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        fallback_path = output_path.with_name(
            f"{output_path.stem}_{stamp}{output_path.suffix}"
        )
        workbook.save(fallback_path)
        return fallback_path


def build_etapa_rows(
    etapa: str,
    item_config: list[dict[str, str | None]],
    source_criteria: dict[str, dict[str, str | None]],
) -> list[list[str | None]]:
    rows = []

    for item in item_config:
        if item["kind"] == "section":
            rows.append(
                build_section_row(
                    item["code"],
                    etapa,
                    None,
                    item["dimensao"],
                    item["obs"],
                )
            )
            continue

        criterio = item.get("source_key", item["criterio"])
        rows.append(
            build_criterion_row(
                item,
                item["code"],
                etapa,
                None,
                item["dimensao"],
                source_criteria[criterio],
            )
        )

    return rows


def style_sheet(ws, header_color: str) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor=header_color)
    body_fill = PatternFill(fill_type="solid", fgColor="F8FBFF")

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    for worksheet_row in ws.iter_rows(min_row=2):
        for cell in worksheet_row:
            cell.fill = body_fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = {
        "A": 12,
        "B": 16,
        "C": 16,
        "D": 18,
        "E": 28,
        "F": 36,
        "G": 28,
        "H": 18,
        "I": 24,
        "J": 28,
        "K": 22,
        "L": 26,
        "M": 28,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    ws.freeze_panes = "A2"


def populate_sheet(ws, headers, rows, header_color: str) -> None:
    ws.append(headers)
    for data_row in rows:
        ws.append(data_row)
    style_sheet(ws, header_color)


def main() -> int:
    source_criteria = load_source_criteria()
    etapa_1_rows = build_etapa_1_rows(source_criteria)
    etapa_2_rows = build_etapa_rows(
        "Etapa 2",
        ETAPA_2_ITEM_CONFIG,
        source_criteria,
    )
    etapa_3_rows = build_etapa_rows(
        "Etapa 3",
        ETAPA_3_ITEM_CONFIG,
        source_criteria,
    )
    synthesis_records = []
    synthesis_records.extend(
        build_criterion_records(
            "Etapa 1",
            ETAPA_1_ITEM_CONFIG,
            source_criteria,
        )
    )
    synthesis_records.extend(
        build_criterion_records(
            "Etapa 2",
            ETAPA_2_ITEM_CONFIG,
            source_criteria,
        )
    )
    synthesis_records.extend(
        build_criterion_records(
            "Etapa 3",
            ETAPA_3_ITEM_CONFIG,
            source_criteria,
        )
    )
    synthesis_rows = build_synthesis_rows(synthesis_records)
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    ws1 = workbook.create_sheet("Etapa 1")
    populate_sheet(ws1, COMMON_HEADERS, etapa_1_rows, "7A3E00")

    ws2 = workbook.create_sheet("Etapa 2")
    populate_sheet(ws2, COMMON_HEADERS, etapa_2_rows, "1F4E78")

    ws3 = workbook.create_sheet("Etapa 3")
    populate_sheet(ws3, COMMON_HEADERS, etapa_3_rows, "38761D")

    ws4 = workbook.create_sheet("Síntese")
    populate_sheet(ws4, SYNTHESIS_HEADERS, synthesis_rows, "5B5B5B")

    output_path = OUTPUT_PATH
    output_path.parent.mkdir(parents=True, exist_ok=True)
    saved_path = save_workbook(workbook, output_path)

    print(f"Arquivo criado: {saved_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
