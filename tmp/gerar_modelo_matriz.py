# Gera o arquivo-modelo da Matriz de Premissas e Critérios (formato congelado)
# a partir da estrutura de data/Matriz_Criterios_Premissas_PLI-SP_FASES_V2_20260715_ATUALIZADA.xlsx
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "data" / "modelos" / "Modelo_Matriz_Criterios_Premissas_SLT.xlsx"
OUT.parent.mkdir(parents=True, exist_ok=True)

HEAD_FILL = PatternFill("solid", fgColor="1F4E5F")
HEAD_FONT = Font(bold=True, color="FFFFFF", size=11)
EX_FILL = PatternFill("solid", fgColor="FFF3CD")
EX_FONT = Font(italic=True, color="7A6000")
THIN = Border(*[Side(style="thin", color="B0B0B0")] * 4)
WRAP = Alignment(wrap_text=True, vertical="top")

wb = Workbook()


def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.border = THIN
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")


def add_sheet(title, headers, widths, examples, freeze="A2", n_blank=200):
    ws = wb.create_sheet(title)
    ws.append(headers)
    style_header(ws, len(headers))
    for ex in examples:
        ws.append(ex)
        r = ws.max_row
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = EX_FILL
            cell.font = EX_FONT
            cell.border = THIN
            cell.alignment = WRAP
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    for r in range(ws.max_row + 1, ws.max_row + 1 + n_blank):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN
            cell.alignment = WRAP
    ws.freeze_panes = freeze
    ws.sheet_properties.tabColor = "1F4E5F"
    return ws


# ---------------- Aba de instruções ----------------
inst = wb.active
inst.title = "Instruções"
inst.sheet_properties.tabColor = "C0392B"
linhas = [
    ("MODELO — MATRIZ DE PREMISSAS E CRITÉRIOS (SLT / PLI-SP)", ""),
    ("", ""),
    ("Como preencher", ""),
    ("1.", "Não altere os nomes das abas nem os cabeçalhos das colunas: o sistema depende deles para interpretar o arquivo."),
    ("2.", "As linhas destacadas em amarelo são EXEMPLOS de preenchimento. Apague-as antes de enviar o arquivo."),
    ("3.", "Preencha a aba 'Dimensões de Critérios' com as dimensões usadas, a aba 'Critérios' com todos os critérios e a aba 'Matriz Crit Premissas v2' com a matriz completa (uma linha por premissa)."),
    ("4.", "A aba 'Índice Risco e Restrição' descreve o índice da Fase 1 (níveis-base, pesos e faixas de classificação)."),
    ("5.", "Coluna 'Fase': use 1 (Elegibilidade territorial), 2 (Favorabilidade territorial) ou 3 (Ajuste por atributos)."),
    ("6.", "Coluna 'Relação': use '↓ Negativa' quando valores maiores pioram a avaliação e '↑ Positiva' quando valores maiores melhoram."),
    ("7.", "Coluna 'Mandatório': use apenas 'Sim' ou 'Não'."),
    ("8.", "Coluna 'Classificação': ex.: 'restrição legal', 'risco inicial / restrição condicionada', 'risco espacial', 'favorabilidade'."),
    ("9.", "Não mescle células, não insira colunas novas e não use fórmulas nas células de dados."),
    ("10.", "Salve o arquivo em formato .xlsx e anexe-o na seção 'Matriz de premissas e critérios' do cadastro da hierarquização."),
    ("", ""),
    ("Abas obrigatórias", "Dimensões de Critérios · Critérios · Matriz Crit Premissas v2 · Índice Risco e Restrição"),
]
for a, b in linhas:
    inst.append([a, b])
inst.column_dimensions["A"].width = 8
inst.column_dimensions["B"].width = 130
inst["A1"].font = Font(bold=True, size=14, color="1F4E5F")
inst["A3"].font = Font(bold=True, size=12)
for r in range(1, inst.max_row + 1):
    inst.cell(row=r, column=2).alignment = WRAP

# ---------------- Dimensões ----------------
add_sheet(
    "Dimensões de Critérios",
    ["#", "Dimensão", "Justificativa (com referência)"],
    [6, 25, 100],
    [[1, "Ambiental", "EXEMPLO — Reúne critérios de restrição e risco ambiental (Unidades de Conservação, APPs, vegetação nativa). Ref.: Lei Federal nº 9.985/2000."]],
)

# ---------------- Critérios ----------------
add_sheet(
    "Critérios",
    ["#", "Critério", "Dimensão"],
    [6, 70, 25],
    [[1, "EXEMPLO — Sobreposição com Unidade de Conservação de Proteção Integral", "Ambiental"]],
)

# ---------------- Matriz ----------------
mat_headers = [
    "Dimensão", "Fase", "Critério", "Premissa", "Classificação", "Dado",
    "Variável", "Unidade de medida", "Operador", "Relação", "Métricas",
    "Fonte", "Mandatório",
]
ws_mat = add_sheet(
    "Matriz Crit Premissas v2",
    mat_headers,
    [16, 8, 45, 60, 30, 45, 45, 26, 45, 14, 45, 45, 12],
    [[
        "Ambiental", 1,
        "EXEMPLO — Sobreposição com Unidade de Conservação de Proteção Integral (estadual)",
        "As unidades estaduais do grupo de Proteção Integral restringem a implantação de empreendimentos.",
        "risco inicial / restrição condicionada",
        "Polígonos oficiais das Unidades de Conservação estaduais",
        "Interseção da demanda com cada feição de UC de Proteção Integral",
        "presença/ausência (0/1)",
        "União temática por Identity, preservando os atributos da UC",
        "↓ Negativa",
        "0 sem interseção; 1 com interseção",
        "DataGeo/IDEA-SP e Fundação Florestal; Lei Federal nº 9.985/2000",
        "Sim",
    ]],
)
dv_fase = DataValidation(type="list", formula1='"1,2,3"', allow_blank=True)
dv_rel = DataValidation(type="list", formula1='"↓ Negativa,↑ Positiva"', allow_blank=True)
dv_mand = DataValidation(type="list", formula1='"Sim,Não"', allow_blank=True)
for dv, col in ((dv_fase, "B"), (dv_rel, "J"), (dv_mand, "M")):
    ws_mat.add_data_validation(dv)
    dv.add(f"{col}2:{col}500")

# ---------------- Índice Fase 1 ----------------
add_sheet(
    "Índice Risco e Restrição",
    ["Ordem", "Critério da Fase 1", "Classificação de origem", "Nível-base",
     "Classe-base", "Peso padrão", "Condição de elevação", "Agregação",
     "Tratamento sem dado", "", "Faixa", "Limite inferior", "Limite superior",
     "Classe arredondada"],
    [8, 45, 30, 12, 16, 12, 45, 26, 18, 4, 14, 14, 14, 18],
    [[1, "EXEMPLO — Sobreposição com UC de Proteção Integral", "risco inicial / restrição condicionada",
      3, "Risco alto", 1, "Elevar para 4 somente com vedação legal expressa",
      "Média ponderada de risco", "Não avaliado", "", "Sem risco", 0, 0, "Sem risco"]],
)

wb.save(OUT)
print("OK:", OUT)
