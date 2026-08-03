"""Reorganiza a aba 'Matriz Crit Premissas v2' do arquivo de matriz de critérios/premissas:

- Renomeia a coluna 'Fase' para 'Etapa'.
- Substitui os valores numéricos 1/2/3 pelos nomes das etapas.
- Move a coluna 'Etapa' para ficar entre 'Premissa' e 'Classificação'.
- Cria uma aba 'Etapas' descrevendo o significado de cada etapa da hierarquização.

Mantém demais abas, estilos, larguras, filtros e freeze panes.
"""

from __future__ import annotations

from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ARQUIVO = Path(
    r"data/Matriz_Criterios_Premissas_PLI-SP_FASES_V2_20260715_ATUALIZADA.xlsx"
)

MAPA_ETAPAS: dict[int, str] = {
    1: "Elegibilidade territorial",
    2: "Favorabilidade territorial e da rede",
    3: "Priorização",
}

DESCRICAO_ETAPAS: list[tuple[int, str, str]] = [
    (
        1,
        "Elegibilidade territorial",
        "Fase de elegibilidade territorial por meio do cruzamento da demanda com as superfícies "
        "de risco e de restrição. Define se o projeto é territorialmente elegível antes de qualquer "
        "análise de mérito.",
    ),
    (
        2,
        "Favorabilidade territorial e da rede",
        "Fase de favorabilidade territorial, por meio do cruzamento da demanda com a superfície de "
        "favorabilidade e com a rede de favorabilidade. Mede o quanto o território e a rede "
        "favorecem a implantação do projeto.",
    ),
    (
        3,
        "Priorização",
        "Etapa de ajuste fino por meio da atribuição de valores aos atributos dos projetos. "
        "Corresponde à priorização final entre projetos já elegíveis e favoráveis.",
    ),
]

# Nova ordem: índices 1-based das colunas atuais na ordem final desejada.
#   Atual: 1 Dimensão | 2 Fase | 3 Critério | 4 Premissa | 5 Classificação | 6..13 restantes
#   Nova : Dimensão | Critério | Premissa | Etapa | Classificação | restantes
NOVA_ORDEM_COLUNAS: list[int] = [1, 3, 4, 2, 5, 6, 7, 8, 9, 10, 11, 12, 13]
IDX_ETAPA_NOVO = NOVA_ORDEM_COLUNAS.index(2) + 1  # posição 1-based da coluna Etapa após reordenar


def copiar_celula(origem, destino) -> None:
    destino.value = origem.value
    if origem.has_style:
        destino.font = copy(origem.font)
        destino.fill = copy(origem.fill)
        destino.border = copy(origem.border)
        destino.alignment = copy(origem.alignment)
        destino.number_format = origem.number_format
        destino.protection = copy(origem.protection)


def reordenar_matriz(wb: openpyxl.Workbook) -> None:
    ws_orig = wb["Matriz Crit Premissas v2"]
    max_row = ws_orig.max_row
    max_col = ws_orig.max_column

    # Cria aba temporária com a nova ordem.
    ws_novo = wb.create_sheet("__matriz_tmp__")

    for nova_col, col_origem in enumerate(NOVA_ORDEM_COLUNAS, start=1):
        for row in range(1, max_row + 1):
            origem = ws_orig.cell(row=row, column=col_origem)
            destino = ws_novo.cell(row=row, column=nova_col)
            copiar_celula(origem, destino)

    # Renomeia cabeçalho 'Fase' -> 'Etapa'.
    ws_novo.cell(row=1, column=IDX_ETAPA_NOVO).value = "Etapa"

    # Substitui valores numéricos da coluna Etapa pelos nomes.
    for row in range(2, max_row + 1):
        celula = ws_novo.cell(row=row, column=IDX_ETAPA_NOVO)
        if celula.value in MAPA_ETAPAS:
            celula.value = MAPA_ETAPAS[celula.value]

    # Reaplica larguras de coluna na nova ordem.
    for nova_col, col_origem in enumerate(NOVA_ORDEM_COLUNAS, start=1):
        letra_origem = get_column_letter(col_origem)
        letra_nova = get_column_letter(nova_col)
        dim_origem = ws_orig.column_dimensions.get(letra_origem)
        if dim_origem is not None and dim_origem.width:
            ws_novo.column_dimensions[letra_nova].width = dim_origem.width

    # A coluna Etapa passa a comportar textos maiores; alarga se necessário.
    letra_etapa = get_column_letter(IDX_ETAPA_NOVO)
    largura_atual = ws_novo.column_dimensions[letra_etapa].width or 0
    if largura_atual < 32:
        ws_novo.column_dimensions[letra_etapa].width = 32

    # Alturas de linha (preserva as definidas).
    for row_idx, dim in ws_orig.row_dimensions.items():
        if dim.height:
            ws_novo.row_dimensions[row_idx].height = dim.height

    ws_novo.freeze_panes = ws_orig.freeze_panes
    ws_novo.sheet_view.showGridLines = ws_orig.sheet_view.showGridLines
    ws_novo.auto_filter.ref = (
        f"A1:{get_column_letter(len(NOVA_ORDEM_COLUNAS))}{max_row}"
    )

    # Substitui a aba original pela nova, mantendo o nome e a posição.
    posicao = wb.sheetnames.index("Matriz Crit Premissas v2")
    del wb["Matriz Crit Premissas v2"]
    ws_novo.title = "Matriz Crit Premissas v2"
    wb.move_sheet(ws_novo, offset=posicao - (len(wb.sheetnames) - 1))


def criar_aba_etapas(wb: openpyxl.Workbook) -> None:
    nome = "Etapas"
    if nome in wb.sheetnames:
        del wb[nome]
    ws = wb.create_sheet(nome, index=0)

    header_fill = PatternFill("solid", fgColor="FF1F4E78")
    header_font = Font(bold=True, color="FFFFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_align = Alignment(vertical="top", wrap_text=True)
    thin = Side(style="thin", color="FFBFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = (
        "Etapas do processo de hierarquização — significado dos valores da coluna "
        "'Etapa' na aba 'Matriz Crit Premissas v2'."
    )
    ws["A1"].font = Font(italic=True, color="FF404040")
    ws.merge_cells("A1:C1")
    ws.row_dimensions[1].height = 32
    ws["A1"].alignment = Alignment(vertical="center", wrap_text=True)

    cabecalhos = ["#", "Etapa", "Descrição"]
    for col, texto in enumerate(cabecalhos, start=1):
        c = ws.cell(row=3, column=col, value=texto)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = border

    for i, (num, nome_etapa, descricao) in enumerate(DESCRICAO_ETAPAS, start=4):
        ws.cell(row=i, column=1, value=num).alignment = Alignment(
            horizontal="center", vertical="top"
        )
        ws.cell(row=i, column=2, value=nome_etapa).alignment = body_align
        ws.cell(row=i, column=3, value=descricao).alignment = body_align
        for col in range(1, 4):
            ws.cell(row=i, column=col).border = border
        ws.row_dimensions[i].height = 90

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 90
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False


def main() -> None:
    caminho = Path(__file__).resolve().parents[1] / ARQUIVO
    print(f"Abrindo: {caminho}")
    wb = openpyxl.load_workbook(caminho)
    reordenar_matriz(wb)
    criar_aba_etapas(wb)
    wb.save(caminho)
    print("Arquivo atualizado.")
    print("Ordem final das abas:", wb.sheetnames)


if __name__ == "__main__":
    main()
